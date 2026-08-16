from __future__ import annotations

import csv
import base64
import hmac
import io
import json
import os
import re
import sqlite3
import time
import urllib.error
import urllib.parse
import urllib.request
from contextlib import contextmanager
from datetime import datetime, timedelta
from pathlib import Path
from threading import Lock, Thread
from typing import Any
from uuid import uuid4
from zipfile import BadZipFile

from flask import Flask, jsonify, request, send_from_directory
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from knowledge_seed import seed_knowledge
from inventory import (
    INVENTORY_SCHEMA,
    InventoryError,
    create_fulfillment_shipment,
    ensure_order_item_from_quote,
    initialize_inventory,
    register_inventory_routes,
)
from operations import OPERATIONS_SCHEMA, register_operations_routes
from rpa.weixin_driver import RPAError, SearchStatus, WeixinDriver

try:
    import win32crypt
except ImportError:  # pragma: no cover - production target is Windows
    win32crypt = None

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data" / "durian_agent.db"
DEFAULT_CONTACT_SOURCE = "公开网站的商户公开页面"
DEFAULT_CONTACT_SOURCE_BASIS = "业务方确认：该批联系人来自公开网站展示的商户公开信息"
MAX_DAILY_FRIEND_ADDS = 150

app = Flask(__name__, static_folder="static", static_url_path="/static")
app.config["JSON_AS_ASCII"] = False
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0


@app.after_request
def prevent_stale_console_assets(response):
    """The console is shipped as mutable static files; never serve a stale UI after an upgrade."""
    if request.path == "/" or request.path.startswith("/static/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@app.before_request
def require_admin_token():
    """Protect central business APIs when the production admin token is configured."""
    if not request.path.startswith("/api/") or request.path in {"/api/health"}:
        return None
    if request.path.startswith("/api/worker/"):
        return None
    expected = os.environ.get("AGENT_ADMIN_TOKEN", "").strip()
    if not expected:
        return None
    supplied = request.headers.get("X-Admin-Token", "").strip()
    if not supplied or not hmac.compare_digest(expected, supplied):
        return jsonify({"error": "中央控制台管理口令无效", "code": "admin_token_required"}), 401
    return None


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@contextmanager
def db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 5000")
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def rows_to_dict(rows):
    return [dict(row) for row in rows]


def audit(conn, action: str, object_type: str, object_id: Any, detail: str = ""):
    conn.execute(
        "INSERT INTO audit_logs(action, object_type, object_id, detail, created_at) VALUES(?,?,?,?,?)",
        (action, object_type, str(object_id), detail, now()),
    )


SCHEMA = """
CREATE TABLE IF NOT EXISTS leads (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    store_name TEXT NOT NULL,
    contact_name TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    wechat_id TEXT DEFAULT '',
    region TEXT DEFAULT '',
    store_type TEXT DEFAULT '',
    source TEXT DEFAULT '',
    source_basis TEXT DEFAULT '',
    import_provenance TEXT DEFAULT '',
    scale TEXT DEFAULT '',
    durian_status TEXT DEFAULT '',
    score INTEGER DEFAULT 50,
    status TEXT DEFAULT '待触达',
    owner TEXT DEFAULT '未分配',
    tags TEXT DEFAULT '',
    next_follow_at TEXT,
    last_contact_at TEXT,
    stop_marketing INTEGER DEFAULT 0,
    notes TEXT DEFAULT '',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);
CREATE UNIQUE INDEX IF NOT EXISTS idx_lead_contact ON leads(phone, wechat_id, store_name);

CREATE TABLE IF NOT EXISTS wechat_accounts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nickname TEXT NOT NULL,
    wechat_no TEXT NOT NULL UNIQUE,
    channel_mode TEXT DEFAULT 'mock',
    approval_ref TEXT DEFAULT '',
    daily_limit INTEGER DEFAULT 150,
    used_today INTEGER DEFAULT 0,
    quota_date TEXT DEFAULT '',
    status TEXT DEFAULT '在线',
    last_sync_at TEXT,
    created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS friend_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lead_id INTEGER NOT NULL,
    account_id INTEGER NOT NULL,
    greeting TEXT NOT NULL,
    remark TEXT DEFAULT '',
    region_snapshot TEXT DEFAULT '',
    status TEXT DEFAULT '待执行',
    result_note TEXT DEFAULT '',
    scheduled_at TEXT,
    executed_at TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY(lead_id) REFERENCES leads(id),
    FOREIGN KEY(account_id) REFERENCES wechat_accounts(id)
);

CREATE TABLE IF NOT EXISTS conversations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lead_id INTEGER NOT NULL UNIQUE,
    stage TEXT DEFAULT '初次沟通',
    intent TEXT DEFAULT '待识别',
    sentiment TEXT DEFAULT '中性',
    ai_mode TEXT DEFAULT '建议回复',
    unread INTEGER DEFAULT 0,
    human_takeover INTEGER DEFAULT 0,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(lead_id) REFERENCES leads(id)
);

CREATE TABLE IF NOT EXISTS messages (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    conversation_id INTEGER NOT NULL,
    sender TEXT NOT NULL,
    content TEXT NOT NULL,
    message_type TEXT DEFAULT 'text',
    created_at TEXT NOT NULL,
    FOREIGN KEY(conversation_id) REFERENCES conversations(id)
);

CREATE TABLE IF NOT EXISTS products (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    sku TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    origin TEXT DEFAULT '',
    grade TEXT DEFAULT '',
    unit TEXT DEFAULT '箱',
    price REAL NOT NULL,
    stock INTEGER DEFAULT 0,
    status TEXT DEFAULT '在售',
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS quotes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    lead_id INTEGER NOT NULL,
    product_id INTEGER NOT NULL,
    quantity INTEGER NOT NULL,
    unit_price REAL NOT NULL,
    freight REAL DEFAULT 0,
    total REAL NOT NULL,
    valid_until TEXT NOT NULL,
    status TEXT DEFAULT '待确认',
    created_at TEXT NOT NULL,
    FOREIGN KEY(lead_id) REFERENCES leads(id),
    FOREIGN KEY(product_id) REFERENCES products(id)
);

CREATE TABLE IF NOT EXISTS orders (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_no TEXT NOT NULL UNIQUE,
    lead_id INTEGER NOT NULL,
    quote_id INTEGER,
    amount REAL NOT NULL,
    payment_status TEXT DEFAULT '待付款',
    status TEXT DEFAULT '待确认',
    receiver TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    address TEXT DEFAULT '',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(lead_id) REFERENCES leads(id),
    FOREIGN KEY(quote_id) REFERENCES quotes(id)
);

CREATE TABLE IF NOT EXISTS shipments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id INTEGER NOT NULL UNIQUE,
    carrier TEXT NOT NULL,
    tracking_no TEXT NOT NULL,
    batch_no TEXT DEFAULT '',
    status TEXT DEFAULT '已揽收',
    shipped_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    FOREIGN KEY(order_id) REFERENCES orders(id)
);

CREATE TABLE IF NOT EXISTS audit_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    action TEXT NOT NULL,
    object_type TEXT NOT NULL,
    object_id TEXT DEFAULT '',
    detail TEXT DEFAULT '',
    created_at TEXT NOT NULL
);
"""


def init_db():
    with db() as conn:
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute("PRAGMA synchronous = NORMAL")
        conn.executescript(SCHEMA)
        conn.executescript(OPERATIONS_SCHEMA)
        conn.executescript(INVENTORY_SCHEMA)
        lead_columns = {row[1] for row in conn.execute("PRAGMA table_info(leads)").fetchall()}
        if "import_provenance" not in lead_columns:
            conn.execute("ALTER TABLE leads ADD COLUMN import_provenance TEXT DEFAULT ''")
        # 文件名、工作表和行号只能用于内部追溯，不能冒充联系人授权来源。
        conn.execute(
            """
            UPDATE leads
            SET import_provenance=source_basis,
                source=?, source_basis=?
            WHERE source IN ('Excel导入','CSV导入')
              AND source_basis LIKE '% / % / 第%行'
              AND COALESCE(import_provenance,'')=''
            """,
            (DEFAULT_CONTACT_SOURCE, DEFAULT_CONTACT_SOURCE_BASIS),
        )
        # 旧版本为联调联系人写入了“微信测试/用户指定进行会话”；按业务方本次确认更正为真实公开来源。
        conn.execute(
            """
            UPDATE leads SET source=?,source_basis=?,updated_at=?
            WHERE source='微信测试'
            """,
            (DEFAULT_CONTACT_SOURCE, DEFAULT_CONTACT_SOURCE_BASIS, now()),
        )
        account_columns = {row[1] for row in conn.execute("PRAGMA table_info(wechat_accounts)").fetchall()}
        if "quota_date" not in account_columns:
            conn.execute("ALTER TABLE wechat_accounts ADD COLUMN quota_date TEXT DEFAULT ''")
        task_columns = {row[1] for row in conn.execute("PRAGMA table_info(friend_tasks)").fetchall()}
        if "remark" not in task_columns:
            conn.execute("ALTER TABLE friend_tasks ADD COLUMN remark TEXT DEFAULT ''")
        if "region_snapshot" not in task_columns:
            conn.execute("ALTER TABLE friend_tasks ADD COLUMN region_snapshot TEXT DEFAULT ''")
        conn.execute(
            """UPDATE friend_tasks
               SET region_snapshot=COALESCE((SELECT region FROM leads WHERE leads.id=friend_tasks.lead_id),'')
               WHERE COALESCE(region_snapshot,'')=''"""
        )
        conn.execute(
            "UPDATE wechat_accounts SET channel_mode='rpa',approval_ref=CASE WHEN approval_ref='人工确认模式' THEN '腾讯RPA授权范围' ELSE approval_ref END WHERE channel_mode='manual'"
        )
        conn.execute(
            """
            UPDATE friend_tasks
            SET status='待执行', result_note='系统重启，任务已恢复到待执行'
            WHERE status='执行中'
            """
        )
        seed_knowledge(conn, now())
        # 微信获客账号不按华南、华东分组；地区跟随每条CSV/Excel线索。
        conn.execute("UPDATE wechat_accounts SET nickname='微信执行账号1' WHERE nickname='华南销售号'")
        conn.execute("UPDATE wechat_accounts SET nickname='微信执行账号2' WHERE nickname='华东销售号'")
        conn.execute("UPDATE wechat_accounts SET daily_limit=?", (MAX_DAILY_FRIEND_ADDS,))
        if conn.execute("SELECT COUNT(*) FROM wechat_accounts").fetchone()[0] == 0:
            conn.executemany(
                "INSERT INTO wechat_accounts(nickname,wechat_no,channel_mode,approval_ref,daily_limit,used_today,quota_date,status,last_sync_at,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)",
                [
                    ("微信执行账号1", "durian_sales_01", "rpa", "腾讯RPA授权范围", MAX_DAILY_FRIEND_ADDS, 0, datetime.now().strftime("%Y-%m-%d"), "在线", now(), now()),
                    ("微信执行账号2", "durian_sales_02", "rpa", "腾讯RPA授权范围", MAX_DAILY_FRIEND_ADDS, 0, datetime.now().strftime("%Y-%m-%d"), "在线", now(), now()),
                ],
            )
        if conn.execute("SELECT COUNT(*) FROM products").fetchone()[0] == 0:
            conn.executemany(
                "INSERT INTO products(sku,name,origin,grade,unit,price,stock,status,updated_at) VALUES(?,?,?,?,?,?,?,?,?)",
                [
                    ("D-MY-MK-A", "马来西亚猫山王", "马来西亚", "A级", "箱", 1280, 86, "在售", now()),
                    ("D-TH-MT-A", "泰国金枕", "泰国", "A级", "箱", 680, 220, "在售", now()),
                    ("D-VN-G6-B", "越南干尧", "越南", "B级", "箱", 520, 145, "在售", now()),
                ],
            )
        initialize_inventory(conn, now())


@app.get("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


@app.get("/api/health")
def health():
    return jsonify({"ok": True, "time": now(), "channel": "rpa"})


def reset_daily_quota(conn):
    today = datetime.now().strftime("%Y-%m-%d")
    conn.execute(
        "UPDATE wechat_accounts SET used_today=0,quota_date=? WHERE COALESCE(quota_date,'')<>?",
        (today, today),
    )


RPA_JOB_LOCK = Lock()
RPA_EXECUTION_LOCK = Lock()
RPA_JOBS: dict[str, dict[str, Any]] = {}
RPA_LATEST_JOB_ID: str | None = None


def render_task_text(template: str, lead: sqlite3.Row, limit: int) -> str:
    replacements = {
        "{店名}": lead["store_name"] or "",
        "{门店}": lead["store_name"] or "",
        "{联系人}": lead["contact_name"] or "",
        "{地区}": lead["region"] or "",
    }
    result = template.strip()
    for key, value in replacements.items():
        result = result.replace(key, value)
    return result[:limit]


KNOWLEDGE_KEYWORDS = [
    "猫山王", "黑刺", "整箱", "单粒", "价格", "多少钱", "报价", "运费", "华南", "常规", "偏远",
    "少房", "虫果", "夹生", "生果", "核大", "口感", "售后", "退款", "赔偿", "物流", "代发",
    "供应商", "太贵", "贵了", "不需要", "别联系", "试单", "产地", "采摘", "库存", "现货", "新鲜", "鲜果",
    "你好", "您好", "哈喽", "在吗", "hello", "hi", "你是谁", "请问你是", "哪位", "你哪位", "身份", "怎么称呼", "如何称呼", "叫什么", "哪个公司", "哪家公司", "公司名称", "做什么的", "干嘛的",
    "联系方式", "号码来源", "信息来源", "怎么有我", "哪里来的", "怎么知道", "谁给的", "通讯录",
    "隐私", "泄露", "买卖信息", "授权", "同意", "删除", "删掉", "举报",
    "付款", "怎么付", "打款", "发票", "开票", "专票", "资质", "营业执照", "图片", "视频", "实拍", "起订量", "最低多少", "多久到", "几天到", "几天能到",
    "成熟度", "甜度", "品质", "稳定", "直接吃", "明天到", "长期合作", "合作价", "忙", "晚点", "已有供应商", "试一粒",
    "机器人", "AI", "ai", "真人", "人工", "自动回复", "系统回复", "销售助手",
    "骗子", "骗局", "诈骗", "靠谱", "可信", "证明你们", "打款安全",
    "包装", "纸箱", "保温", "破损", "拒收", "漏液", "压坏", "催单", "快递", "物流单号", "到哪了",
    "优惠", "便宜点", "量大", "批量价", "利润", "赚多少", "好卖", "合同", "账期", "长期价", "改地址", "取消订单",
    "复购", "回购", "老客户", "再来", "态度", "投诉", "骂人", "看不懂", "什么意思",
]

UNKNOWN_FACT_RISK_TERMS = (
    "价格", "多少钱", "报价", "运费", "库存", "现货", "到货", "发货", "时效",
    "产地", "证明", "退款", "赔偿", "赔付", "售后", "合同", "账期", "付款", "订单", "发票", "开票", "专票",
    "联系方式", "号码来源", "信息来源", "隐私", "授权", "同意", "泄露", "买卖信息",
    "包装", "打款", "优惠", "批量价", "合同", "账期", "改地址", "取消订单", "物流单号",
    "到哪了", "破损", "拒收", "漏液", "压坏", "证明你们", "营业执照",
)

AUTOMATION_IDENTITY_TERMS = (
    "机器人", "ai吗", "ai回复", "人工智能", "自动回复", "系统回复", "是真人吗", "是不是本人",
    "你是本人吗", "是真人不", "是人工吗", "人工回复吗", "你是真人", "你是活人", "有人在看吗",
)

COMMON_SCENE_RESPONSES = (
    (("骗子", "骗局", "诈骗", "靠不靠谱", "正规吗", "可信不", "打款安全吗"),
     "信任与防诈说明-01",
     "您谨慎一点是对的。公司和产品信息我会按已经核验的内容说明，涉及收款、合同或资质材料时，只使用公司确认过的信息；没有核验清楚前，您先不要付款。",
     "信任与防诈边界"),
    (("包装怎么样", "怎么包装", "什么包装", "纸箱", "保温箱", "会不会压坏"),
     "包装方式待确认",
     "包装要看本次发货方式和当批规格，我这里不能先说死。您告诉我整箱还是单粒、发到哪里，我按实际包装方案核实。",
     "包装方案待核实"),
    (("甜不甜", "好不好吃", "口感怎么样", "甜度多少", "保证甜", "直接吃", "每颗都稳定", "成熟度一样"),
     "口感成熟度边界",
     "榴莲属于生鲜，不同批次和成熟度会有差异，我不能保证每一颗甜度完全一样。可以按品种、等级和当批果况给您说明，但个人口感不能作统一承诺。",
     "口感与成熟度边界"),
    (("便宜点", "还能优惠", "有优惠吗", "量大什么价", "批量价", "最低价", "底价"),
     "优惠价格待核",
     "可以按品种、规格、数量和收货地区一起核，但我不能先承诺一个固定最低价。您把预计数量发我，我按当前规则确认。",
     "优惠价格待核实"),
    (("利润多少", "能赚多少", "好不好卖", "好卖吗", "利润怎么样", "毛利多少"),
     "门店利润边界",
     "门店实际利润还要看您的零售价、损耗、运费和当地客群，我不能替您保证销量或利润。您给我常卖规格和零售价，我可以帮您按进货成本把账算清楚。",
     "门店利润测算边界"),
    (("签合同", "合同怎么签", "能签合同", "账期", "月结", "先货后款"),
     "合同账期待确认",
     "合同和账期要按公司当前政策、合作主体和订单金额确认，我不能在聊天里直接承诺。您把公司主体和预计采购量发我，由负责人核实。",
     "合同账期需核实"),
    (("改地址", "地址写错", "换地址", "取消订单", "不要了", "撤单"),
     "订单变更待确认",
     "我先帮您暂停继续处理。地址能否修改或订单能否取消，要看当前是否已经占库、付款或发货；请把订单号和需要变更的内容发我核实。",
     "订单变更需核实"),
    (("物流单号", "快递单号", "到哪了", "怎么还没到", "催一下物流", "物流没更新"),
     "物流查询待核实",
     "可以查，但需要对应订单号或运单号。我不能凭聊天记录猜物流状态，您把编号发我，我按实际物流记录核对。",
     "物流状态需核实"),
    (("包装破", "破损", "漏液", "压坏", "拒收", "外箱坏", "外箱破"),
     "运输异常处理",
     "先别丢外箱和面单，麻烦把外包装、面单和果况拍清楚；是否拒收、补发或赔付要结合物流记录和实际损坏情况核实。",
     "运输异常需核实"),
    (("再来一", "回购", "复购", "上次那种", "老客户", "还是原来", "按原来"),
     "复购确认-01",
     "可以老板。复购也要重新确认当期价格、库存和收货信息，您先说这次还是原来的品种规格和数量吗？",
     "复购信息确认"),
    (("看不懂", "什么意思", "没明白", "说简单点", "太复杂"),
     "简化说明-01",
     "可以，我说简单一点。您现在最想先确认哪一件事，我只把这一项说清楚。",
     "简化表达"),
    (("态度不好", "你什么态度", "别废话", "烦不烦", "说人话"),
     "情绪降级-01",
     "抱歉，刚才说得不合适。我不绕了，您把现在最需要确认的问题发我，我直接回答。",
     "负面情绪降级"),
)

CONTACT_SOURCE_TERMS = (
    "怎么获得我联系方式", "怎么得到我联系方式", "怎么有我的联系方式", "联系方式哪来的",
    "联系方式哪里来的", "号码哪来的", "号码哪里来的", "怎么有我号码", "怎么知道我电话",
    "谁给你的号码", "谁给的联系方式", "从哪里找到我", "从哪找到我", "为什么加我",
    "怎么知道我是水果店", "我的信息哪里看到", "哪个渠道找到我", "从哪个渠道找到我",
    "谁把我联系方式给你", "谁把我的联系方式给你", "谁把我号码给你", "信息从哪里看到",
)
SOURCE_DENIAL_TERMS = (
    "我没登记", "我没有登记", "我没参加", "我没有参加", "我不认识", "我没授权",
    "我没有授权", "我没同意", "我没有同意", "没在网站登记", "没有在网站登记",
    "记录不对", "来源不对", "你说的不对",
)
PRIVACY_COMPLAINT_TERMS = (
    "信息泄露", "泄露隐私", "侵犯隐私", "非法获取", "买卖信息", "买的名单", "举报你们", "投诉你们",
)
STOP_CONTACT_TERMS = (
    "不要联系", "不要再联系", "别联系", "别再联系", "不再联系", "别再发", "不要再发", "停止联系",
    "把我删了", "删掉我", "删除我的联系方式", "删除我信息", "删除我的信息", "不需要了", "别打扰",
)


def is_contact_source_question(text: str) -> bool:
    normalized = re.sub(r"\s+", "", text or "")
    if any(term in normalized for term in CONTACT_SOURCE_TERMS):
        return True
    mentions_contact = any(term in normalized for term in ("联系方式", "号码", "电话", "我的信息", "我是水果店"))
    asks_origin = any(term in normalized for term in (
        "怎么", "哪来", "哪里", "来源", "获得", "得到", "知道", "谁给", "谁把", "渠道", "看到",
    ))
    return mentions_contact and asks_origin


def is_source_denial(text: str) -> bool:
    return any(term in (text or "") for term in SOURCE_DENIAL_TERMS)


def is_privacy_complaint(text: str) -> bool:
    return any(term in (text or "") for term in PRIVACY_COMPLAINT_TERMS)


def is_stop_contact_request(text: str) -> bool:
    return any(term in (text or "") for term in STOP_CONTACT_TERMS)


def is_automation_identity_question(text: str) -> bool:
    normalized = re.sub(r"\s+", "", (text or "").lower())
    if any(term in normalized for term in AUTOMATION_IDENTITY_TERMS):
        return True
    if any(term in normalized for term in ("机器人", "人工智能", "自动回复", "系统回复")):
        return True
    if re.search(r"(^|[^a-z])ai([^a-z]|$)", normalized):
        return True
    return "真人" in normalized and any(term in normalized for term in ("吗", "还是", "是不是", "到底", "真不真"))


def decision_requires_human(decision_basis: str) -> bool:
    return decision_basis in {
        "隐私投诉需核查", "来源记录发生冲突", "停止联系请求", "信任与防诈边界",
        "包装方案待核实", "优惠价格待核实", "合同账期需核实", "订单变更需核实",
        "物流状态需核实", "运输异常需核实", "开票规则待核实", "库存需人工核实",
        "付款方式需核实", "资质材料需核实", "批次资料需核实", "起订量需核实",
        "必须人工确认",
    }


def contact_source_is_verified(source: str | None, source_basis: str | None) -> bool:
    source_text = (source or "").strip()
    basis_text = (source_basis or "").strip()
    placeholders = {"", "待核验", "未知", "未填写", "Excel导入", "CSV导入", "表格导入"}
    if source_text in placeholders or basis_text in placeholders:
        return False
    if re.search(r"\s/\s.+\s/\s第\d+行$", basis_text):
        return False
    return True


def unknown_fact_requires_human(text: str) -> bool:
    return any(term in text for term in UNKNOWN_FACT_RISK_TERMS)


def contextual_reply_has_unsupported_fact(text: str) -> bool:
    if re.search(r"\d+(?:\.\d+)?\s*元", text):
        return True
    forbidden = (
        "现货充足", "保证新鲜", "保证到货", "包赔", "全额退款",
        "冻肉", "冻品", "冷冻果肉", "加工品", "最后几箱",
    )
    return any(term in text for term in forbidden)


def reply_contains_false_human_claim(text: str) -> bool:
    forbidden = ("我是真人", "我不是机器人", "本人纯手工回复", "全程人工本人", "真人一对一回复")
    return any(term in (text or "") for term in forbidden)


def knowledge_risk_level(matches: list[dict[str, Any]], requires_human: bool) -> str:
    if requires_human or any(row.get("risk_level") == "高" for row in matches):
        return "高"
    if any(row.get("risk_level") == "中" for row in matches):
        return "中"
    return "低"


def auto_label_knowledge(content: str, requested_category: str = "") -> dict[str, str]:
    text = content.strip()
    if requested_category:
        category = requested_category
    elif any(k in text for k in ("多少钱", "价格", "报价", "运费")):
        category = "价格话术"
    elif any(k in text for k in ("少房", "虫果", "夹生", "生果", "退款", "赔偿", "售后")):
        category = "售后规则"
    elif any(k in text for k in ("不需要", "别联系", "不再跟进")):
        category = "销售话术"
    else:
        category = "销售话术"
    risk = "高" if any(k in text for k in ("退款", "赔偿", "全退", "全补", "库存", "产地", "手工采摘")) else "低"
    automation = "必须人工确认" if risk == "高" else "可自动建议"
    tags = ",".join(k for k in KNOWLEDGE_KEYWORDS if k in text)
    return {"category": category, "risk_level": risk, "automation_level": automation, "tags": tags}


def search_knowledge_rows(conn, query: str, category: str = "", limit: int = 20) -> list[dict[str, Any]]:
    sql = "SELECT * FROM knowledge_entries WHERE status='已发布'"
    params: list[Any] = []
    if category:
        sql += " AND category=?"
        params.append(category)
    rows = rows_to_dict(conn.execute(sql, params).fetchall())
    query = query.strip().lower()
    if not query:
        return rows[:limit]
    terms = [k.lower() for k in KNOWLEDGE_KEYWORDS if k.lower() in query]
    acknowledgement_words = {"ok", "okay", "好的", "好", "嗯", "嗯嗯", "可以", "行", "收到", "明白", "1"}
    query_lines = {line.strip(" \t，。,.!！?？").lower() for line in query.splitlines() if line.strip()}
    if query_lines.intersection(acknowledgement_words):
        terms.extend(["简短回应推进", "连续推进", "语气适配"])
    if query in {"你好", "您好", "哈喽", "hello", "hi"}:
        terms.extend(["首次开场", "店铺介绍"])
    if any(term in query for term in ("这次对话结束", "结束对话", "今天先聊到这里", "先聊到这里", "不聊了")):
        terms.extend(["对话结束"])
    terms.extend(token.lower() for token in re.findall(r"[A-Za-z0-9.\-]{2,}", query))
    terms.append(query)
    scored = []
    for row in rows:
        haystack = " ".join(str(row.get(key, "")) for key in ("title", "content", "tags", "category")).lower()
        score = sum(8 if term == query else 2 for term in set(terms) if term and term in haystack)
        if score:
            row["score"] = score
            try:
                row["structured"] = json.loads(row.pop("structured_json") or "{}")
            except json.JSONDecodeError:
                row["structured"] = {}
            scored.append(row)
    scored.sort(key=lambda item: (-item["score"], item["risk_level"] == "高", item["id"]))
    return scored[:limit]


def calculate_knowledge_price(conn, payload: dict[str, Any]) -> dict[str, Any]:
    mode = (payload.get("sale_mode") or "").strip()
    product = (payload.get("product") or "").strip()
    grade = (payload.get("grade") or "").strip()
    spec = (payload.get("spec") or "").strip()
    region_group = (payload.get("region_group") or "").strip()
    if not all((mode, product, spec, region_group)):
        raise ValueError("需要明确销售方式、商品、规格和地区档位")
    price = conn.execute(
        "SELECT * FROM price_rules WHERE sale_mode=? AND product=? AND grade=? AND spec=? AND status='启用'",
        (mode, product, grade, spec),
    ).fetchone()
    if not price:
        raise ValueError("没有匹配的启用价格，不能推测报价")
    freight = conn.execute(
        "SELECT * FROM freight_rules WHERE region_group=? AND sale_mode=? AND status='启用'",
        (region_group, mode),
    ).fetchone()
    if not freight:
        raise ValueError("没有匹配的运费规则，不能推测运费")
    quantity = max(1, int(payload.get("quantity", 1)))
    goods_total = round(float(price["price"]) * quantity, 2)
    freight_total = round(float(freight["price"]) * quantity, 2)
    return {
        "sale_mode": mode, "product": product, "grade": grade, "spec": spec,
        "quantity": quantity, "unit_price": price["price"], "goods_total": goods_total,
        "region_group": region_group, "freight": freight["price"], "freight_total": freight_total,
        "total": round(goods_total + freight_total, 2), "currency": price["currency"],
        "boundary": "报价仅依据当前启用规则；库存、具体省份归类和最终发货时间仍需人工确认。",
    }


def protect_api_key(value: str) -> str:
    if not value:
        return ""
    if win32crypt is None:
        raise RuntimeError("当前环境不能安全保存API密钥")
    protected_result = win32crypt.CryptProtectData(value.encode("utf-8"), "DurianAgentDeepSeek", None, None, None, 0)
    protected = protected_result[1] if isinstance(protected_result, tuple) else protected_result
    return base64.b64encode(protected).decode("ascii")


def reveal_api_key(value: str) -> str:
    if not value or win32crypt is None:
        return ""
    try:
        raw = base64.b64decode(value.encode("ascii"))
        return win32crypt.CryptUnprotectData(raw, None, None, None, 0)[1].decode("utf-8")
    except Exception:
        return ""


def validate_deepseek_base_url(value: str) -> str:
    base_url = value.strip().rstrip("/")
    parsed = urllib.parse.urlparse(base_url)
    if parsed.scheme != "https" or parsed.hostname != "api.deepseek.com":
        raise ValueError("为防止密钥泄露，DeepSeek地址必须是 https://api.deepseek.com")
    return base_url


def deepseek_request(base_url: str, api_key: str, path: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    url = validate_deepseek_base_url(base_url) + path
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    request_obj = urllib.request.Request(url, data=data, headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}, method="POST" if data else "GET")
    try:
        with urllib.request.urlopen(request_obj, timeout=15) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        try:
            detail = json.loads(exc.read().decode("utf-8")).get("error", {}).get("message", "")
        except Exception:
            detail = ""
        raise RuntimeError(f"DeepSeek接口返回{exc.code}：{detail or '请检查密钥、余额和模型'}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError("无法连接DeepSeek，请检查网络") from exc


def build_answer_context(conn, question: str, matches: list[dict[str, Any]]) -> tuple[str, bool, list[str]]:
    contexts = [f"[{row['title']}] {row['content']}" for row in matches[:8]]
    requires_human = any(row.get("automation_level") == "必须人工确认" or row.get("risk_level") == "高" for row in matches[:8])
    sources = [row["title"] for row in matches[:8]]
    if any(term in question for term in ("价格", "多少钱", "报价", "到手价", "运费", "整箱", "单粒", "下单", "要一", "来一")):
        price_rows = conn.execute("SELECT * FROM price_rules WHERE status='启用'").fetchall()
        scored_prices = []
        for row in price_rows:
            score = 0
            for field, weight in (("sale_mode", 3), ("product", 4), ("grade", 2), ("spec", 5)):
                value = str(row[field] or "")
                if value and value in question:
                    score += weight
            if score:
                scored_prices.append((score, row))
        for _, row in sorted(scored_prices, key=lambda item: -item[0])[:12]:
            grade = row["grade"] or "无等级"
            contexts.append(f"[启用价格规则] {row['sale_mode']}，{row['product']}，{grade}，{row['spec']}：{row['price']}元")
            sources.append(f"价格规则-{row['sale_mode']}-{row['product']}-{grade}-{row['spec']}")
        freight_rows = conn.execute("SELECT * FROM freight_rules WHERE status='启用' ORDER BY region_group,sale_mode").fetchall()
        for row in freight_rows:
            if row["region_group"] in question or any(term in question for term in ("运费", "到手价", "下单")):
                contexts.append(f"[启用运费规则] {row['region_group']}，{row['sale_mode']}：{row['price']}元")
                sources.append(f"运费规则-{row['region_group']}-{row['sale_mode']}")
    for issue in ("少房", "虫果", "夹生", "生果", "个人口感", "核大"):
        if issue in question:
            rules = conn.execute("SELECT * FROM aftersales_rules WHERE issue=? AND status='启用'", (issue,)).fetchall()
            for rule in rules:
                contexts.append(f"[售后规则] {rule['issue']}，条件：{rule['condition_text']}，处理：{rule['resolution']}。边界：{rule['boundary_note']}")
                sources.append(f"售后规则-{rule['issue']}-{rule['condition_text']}")
            requires_human = True
    return "\n".join(contexts), requires_human, list(dict.fromkeys(sources))


def local_knowledge_answer(
    conn,
    question: str,
    matches: list[dict[str, Any]],
    conversation_context: str = "",
    lead_context: dict[str, Any] | sqlite3.Row | None = None,
) -> tuple[str, str]:
    combined = (conversation_context + "\n" + question).strip()
    customer_lines = re.findall(r"^客户：(.*)$", conversation_context, flags=re.MULTILINE)
    customer_turns = customer_lines + [question]
    customer_context = "\n".join(customer_turns) if customer_lines else question

    def entry(title: str, fallback: str) -> str:
        row = conn.execute(
            "SELECT content FROM knowledge_entries WHERE title=? AND status='已发布'",
            (title,),
        ).fetchone()
        return row["content"] if row else fallback

    def yuan(value: Any) -> str:
        amount = float(value)
        return str(int(amount)) if amount.is_integer() else str(amount)

    def latest_choice(choices: tuple[str, ...]) -> str:
        """Return the last-mentioned choice so customer corrections override old turns."""
        for turn in reversed(customer_turns):
            positioned = [(turn.rfind(choice), choice) for choice in choices if choice in turn]
            if positioned:
                return max(positioned)[1]
        return ""

    def latest_regex(pattern: str, flags: int = 0) -> re.Match[str] | None:
        for turn in reversed(customer_turns):
            found = list(re.finditer(pattern, turn, flags=flags))
            if found:
                return found[-1]
        return None

    def detected_sale_mode() -> str:
        for turn in reversed(customer_turns):
            positions = [
                (turn.rfind("整箱"), "整箱"),
                (turn.rfind("单粒"), "单粒"),
                (turn.rfind("一件代发"), "单粒"),
            ]
            quantity_matches = list(re.finditer(r"\d+\s*(箱|粒)", turn))
            if quantity_matches:
                last_quantity = quantity_matches[-1]
                positions.append((last_quantity.start(), "整箱" if last_quantity.group(1) == "箱" else "单粒"))
            valid = [item for item in positions if item[0] >= 0]
            if valid:
                return max(valid)[1]
        return ""

    def next_sales_step() -> tuple[str, str]:
        mode = detected_sale_mode()
        if not mode:
            return "好老板，那咱们接着来。您这次想看整箱还是单粒？", "简短回应推进"
        product = latest_choice(("猫山王", "黑刺"))
        box_products = tuple(
            row["product"] for row in conn.execute(
                "SELECT product FROM price_rules WHERE sale_mode='整箱' AND status='启用' ORDER BY id"
            ).fetchall()
        )
        box_code = latest_choice(box_products)
        if not product and not box_code:
            return "可以，您想先看猫山王还是黑刺？", "简短回应推进"
        single_spec = latest_regex(r"\d+(?:\.\d+)?-\d+(?:\.\d+)?斤")
        if mode == "单粒" and not single_spec:
            return "行，您平时哪个重量段走得比较多？", "简短回应推进"
        if mode == "整箱" and not box_code:
            return "明白。整箱有几档结算规格，您想先看哪一档？", "简短回应推进"
        region = latest_choice(("华南", "常规", "偏远"))
        if not region:
            return "明白。收货在哪个地区？我把运费也一起算上。", "简短回应推进"
        quantity_match = latest_regex(r"(\d+)\s*(箱|粒)")
        if not quantity_match:
            unit = "箱" if mode == "整箱" else "粒"
            return f"好，您准备先拿多少{unit}？我把金额给您核一下。", "简短回应推进"
        spec_text = box_code or (single_spec.group(0) if single_spec else "")
        product_text = box_code or product
        return (
            f"好老板，我记下了：{product_text}{'，' + spec_text if spec_text and spec_text != product_text else ''}，"
            f"数量{quantity_match.group(1)}{quantity_match.group(2)}，发往{region}。我按当前价格和运费给您做最后确认。",
            "成交前信息复述",
        )

    normalized = question.strip(" \t，。,.!！?？").lower()
    if normalized in {"你好", "您好", "哈喽", "hello", "hi", "在吗", "老板在吗", "你好，打扰了"}:
        return entry(
            "首次开场-01",
            "老板您好，我们这里是做榴莲批发的，全都是马来西亚采摘鲜果空运。看到您也是开水果店的，想和您交流合作一下。🤝😊",
        ), "首次开场"
    if normalized in {
        "ok", "okay", "好的", "好", "嗯", "嗯嗯", "可以", "行", "收到", "明白", "1",
        "行继续", "行，继续", "好继续", "好，继续", "可以继续", "可以，继续",
        "知道了继续说", "知道了，继续说", "明白了继续", "明白了，继续",
    }:
        return next_sales_step()

    if is_stop_contact_request(question):
        return entry(
            "删除停止联系-01",
            "收到，我现在为您标记不再联系，后续不再发送营销信息。给您造成打扰，抱歉。",
        ), "停止联系请求"

    if is_privacy_complaint(question):
        return entry(
            "隐私质疑-01",
            "您的质疑我已记录。在来源核清前我们不会继续营销，我也不会在没有证据的情况下判断是否存在泄露或买卖信息；这件事需要负责人核查并向您说明。",
        ), "隐私投诉需核查"

    if is_source_denial(question):
        return entry(
            "客户否认来源-01",
            "收到，既然您反馈来源记录与实际情况不一致，我先暂停联系，不继续推销，并交由负责人核查。如果您要求不再联系，我们会立即标记停止联系。",
        ), "来源记录发生冲突"

    if is_contact_source_question(question):
        source = (lead_context["source"] if lead_context and "source" in lead_context.keys() else "") or ""
        source_basis = (lead_context["source_basis"] if lead_context and "source_basis" in lead_context.keys() else "") or ""
        if contact_source_is_verified(source, source_basis):
            answer = (
                f"老板，您的联系方式来自“{source.strip()}”，我们保存的登记依据是“{source_basis.strip()}”。"
                "我只能按这条记录如实说明。如果记录与实际不符，我先暂停联系并请负责人核查；"
                "如果您不希望再联系，我可以立即为您标记不再联系。"
            )
            basis = "已核验联系人来源记录"
        else:
            answer = (
                f"老板，您的联系方式是我们从{DEFAULT_CONTACT_SOURCE}看到的，主要用于联系水果门店沟通供货合作。"
                "如果您不希望再联系，直接告诉我即可，我会立即为您标记不再联系。"
            )
            basis = "业务方确认的公开信息来源"
        if is_automation_identity_question(question):
            answer += "另外，这个账号由公司销售人员和服务系统共同维护，重要业务结论由负责人核实。"
            basis += "；账号维护方式如实说明"
        if any(term in question for term in ("你是谁", "哪家公司", "哪个公司", "做什么的", "干嘛的")):
            answer = "这里是跨境云有限公司的榴莲供货对接窗口。" + answer
            basis += "；公司身份信息"
        return answer, basis

    if is_automation_identity_question(question):
        return entry(
            "账号维护说明-01",
            "这个账号由跨境云有限公司的销售人员和服务系统共同维护，日常品种、规格和价格咨询会先在这里回复；"
            "涉及库存、付款、合同、订单变更和售后结论时，由公司负责人核实确认。",
        ), "账号维护方式如实说明"

    company_identity_terms = (
        "哪个公司", "哪家公司", "你们公司", "公司名称", "公司叫什么",
        "你们是做什么", "你们是干嘛", "你们做什么", "你们干嘛",
    )
    naming_terms = ("怎么称呼", "如何称呼", "怎么叫你", "叫你什么", "你叫什么")
    identity_terms = (
        "你是谁", "请问你是", "你哪位", "您哪位", "哪位", "你是做什么",
        "你是干嘛", "做啥的", "你是做榴莲", "你们做榴莲", "你们是做榴莲",
    )
    if any(term in normalized for term in company_identity_terms):
        return entry(
            "身份介绍-公司",
            "我们是跨境云有限公司，主要做马来西亚榴莲供货，经营猫山王、黑刺整箱批发和单粒代发。",
        ), "身份信息"
    if any(term in normalized for term in naming_terms):
        return entry(
            "身份介绍-称呼",
            "老板，您叫我小刘就行，我负责这边的榴莲供货对接。",
        ), "身份信息"
    if any(term in normalized for term in identity_terms):
        return entry(
            "身份介绍-业务窗口",
            "老板您好，这里是跨境云有限公司的榴莲供货对接窗口，您叫我小刘就行。"
            "我们主要做马来西亚猫山王、黑刺，支持整箱批发和单粒代发。",
        ), "身份信息"

    confirmed_aftersales: list[str] = []
    if "少一房" in question:
        confirmed_aftersales.append("少一房补一房。")
    if "少两房" in question:
        confirmed_aftersales.append("少两房退该单粒商品款，原运费不退；实际退款按订单里的单粒价格核对。")
    if "虫果" in question and any(term in question for term in ("一房", "1房")):
        confirmed_aftersales.append("一房虫果补一房。")
    if "虫果" in question and any(term in question for term in ("两房", "2房")):
        confirmed_aftersales.append("两房虫果按补整果处理。")
    if "虫果" in question and any(term in question for term in ("三房", "3房", "三房及以上", "3房及以上")):
        confirmed_aftersales.append("三房及以上虫果按补整果处理。")
    if "夹生" in question and any(term in question for term in ("低于一半", "不到一半", "不足一半", "一半以下")):
        confirmed_aftersales.append("夹生比例低于一半时，有多少售后多少，即按核实后的实际夹生比例处理。")
    if "夹生" in question and any(term in question for term in ("达到一半", "超过一半", "一半以上", "一半")) and not any(
        term in question for term in ("低于一半", "不到一半", "不足一半", "一半以下")
    ):
        confirmed_aftersales.append("夹生比例达到一半时补整果。")
    if "生果" in question and "夹生" not in question:
        if any(term in question for term in ("保证", "会不会", "没有", "绝对")):
            confirmed_aftersales.append("生鲜果不能承诺每颗绝对不会出现生果；如果结合实际果况确认属于生果，按全补处理。")
        else:
            confirmed_aftersales.append("确认属于生果的，按全补处理。")
    if "核大" in question:
        confirmed_aftersales.append("确认核大后，可以协调一房售后处理。")
    if any(term in question for term in ("售后时限", "多久内申请", "多久申请", "几小时内", "几天内")):
        confirmed_aftersales.append("目前还没有确认具体售后申请时限，不能自行说成多少小时或多少天，以后续正式标准为准。")
    evidence_question = any(term in question for term in ("售后证据", "需要什么证据", "开箱视频", "快递面单", "榴莲头"))
    if evidence_question:
        confirmed_aftersales.append(
            "售后不要求必须提供完整开箱视频或快递面单，照片不要求拍全套，但必须把榴莲头拍清楚；"
            "每次供货对应的售后标准以当次发布内容为准。"
        )
    logistics_liability = any(term in question for term in ("物流挤压", "运输挤压", "物流破损", "运输破损", "丢件"))
    if logistics_liability:
        confirmed_aftersales.append(
            "物流运输中发生挤压、破损或丢件，由我们承担相应责任；具体补发、退款或赔付方式要结合订单和物流记录核实。"
        )
    if confirmed_aftersales:
        unique_answers = list(dict.fromkeys(confirmed_aftersales))
        if len(unique_answers) == 1:
            return unique_answers[0], "必须人工确认"
        return "\n".join(f"{index}. {answer}" for index, answer in enumerate(unique_answers, start=1)), "必须人工确认"

    for terms, title, fallback, basis in COMMON_SCENE_RESPONSES:
        if any(term in question for term in terms):
            return entry(title, fallback), basis

    if "夹生" in question:
        return (
            "榴莲是生鲜，成熟度会有批次差异，不能承诺每颗到货状态完全一致。"
            "如果实际开果出现夹生，请保留果壳、果肉、面单和开果记录，再按夹生比例交由售后人员人工核实。",
            "必须人工确认",
        )

    if any(k in question for k in ("怎么付款", "付款方式", "怎么打款", "打款方式", "收款账户", "收款码")):
        return entry(
            "付款方式待确认",
            "付款时只使用公司本次订单确认的收款信息，不要向聊天里临时出现的陌生账户转账。您确定订单后，我按公司当前可核验的付款信息给您确认。",
        ), "付款方式需核实"

    if any(k in question for k in ("营业执照", "公司资质", "资质证明", "能证明你们")):
        return entry(
            "公司资质待确认",
            "可以核验，但我不会在资料未确认前随便发一份。您需要核验哪项主体或资质，我按公司留存的有效材料确认。",
        ), "资质材料需核实"

    if any(k in question for k in ("今天的实拍", "当天实拍", "发实拍", "批次图片", "批次视频", "果子视频", "果子图片")):
        return entry(
            "批次图片视频待确认",
            "可以先确认您要看的品种和规格，我再按当天实际批次核对有没有对应图片或视频；不会拿旧图冒充当天实拍。",
        ), "批次资料需核实"

    if any(k in question for k in ("最低多少起订", "多少起订", "起订量", "最低起订", "几粒起订", "几箱起订")):
        return entry(
            "起订量待确认",
            "起订量要按整箱还是单粒、品种规格和当前发货安排确认，我不先随口定一个数。您先说准备怎么拿货，我按当期规则核实。",
        ), "起订量需核实"

    if "一件代发" in question:
        return entry(
            "物流一件代发",
            "支持一件代发。单粒价格按品种、等级和重量段计算，运费按收货地区档位计算；您把规格和地区发我，我给您按规则核算。",
        ), "一件代发规则"

    if any(k in question for k in ("按斤还是按粒", "按斤算", "怎么计价", "计价单位")):
        return (
            "整箱按箱报价，单粒按粒报价；单粒里的重量段是规格区间，不是收到后再按斤临时计价。",
            "计价单位说明",
        )

    if any(k in question for k in ("个人觉得不好吃", "个人口感", "不合口味", "不喜欢这个味")):
        return entry(
            "个人口感售后边界",
            "个人口感差异不作为统一售后依据；如果同时存在少房、虫果、夹生或生果，请把实际果况拍清楚，再按对应售后规则交由售后人员人工核实。",
        ), "必须人工确认"

    if any(k in question for k in ("当天采的吗", "都是当天采", "每天采摘", "手工采摘吗")):
        return entry(
            "产地与采摘说法",
            "业务资料记录为马来西亚榴莲园、每天手工采摘。具体到某一批次的采摘日期和凭证，需要按当批资料核实后再确认。",
        ), "资质材料需核实"

    if any(k in question for k in ("可以交流", "交流一下", "没问题", "可以聊", "好 可以", "好的可以")):
        return entry("自然承接-01", "好啊老板。您平时是整箱备货，还是有订单后单粒代发？我先按您的做法聊。"), "自然承接"

    if any(k in question for k in ("多久到", "几天到", "几天能到", "物流时效", "什么时候到", "什么时候能收到", "明天到")):
        return entry(
            "物流时效-自然回答",
            "常规物流预计2至3天，具体要看揽收时间、收货地区和运输情况，我不能承诺固定日期必达。",
        ), "物流时效边界"

    if any(k in question for k in ("太贵", "贵了", "价格贵", "价太高")):
        return entry(
            "价格异议承接",
            "理解老板，门店要算实际利润。咱们先按同品种、同等级、同重量和运费一起比较，您把正在对比的规格发我，我按同一口径给您看。",
        ), "价格异议承接"

    if any(k in question for k in ("已有供应商", "固定供应商", "固定渠道", "有供应商了", "暂时不换")):
        return entry(
            "已有供应商承接",
            "明白老板，有稳定供应商很正常。我这边先作为备用渠道，后面需要比价、补单或单粒代发时再联系，不催您更换。",
        ), "已有供应商承接"

    if any(k in question for k in ("试一粒", "先试", "不想拿整箱", "试单")):
        return entry(
            "试单承接",
            "第一次合作可以先从单粒试单，不用直接整箱压货。您告诉我想看的品种、重量段和收货地区，我按现有价格和运费给您核算。",
        ), "试单推进"

    if any(k in question for k in ("现在忙", "在忙", "晚点说", "晚点再聊", "稍后聊")):
        return entry(
            "客户正在忙",
            "好的老板，您先忙，我不连续发消息。您方便的时候回我就行。",
        ), "客户要求稍后沟通"

    if any(k in question for k in ("发票", "开票", "专票", "普票")):
        return entry(
            "发票规则待确认",
            "当前资料里没有确认开票类型、税率和抬头要求，我先不随口答应。您把需要的发票类型告诉我，我让负责人核实后回复。",
        ), "开票规则待核实"

    wants_price = any(k in question for k in ("多少钱", "价格", "报价", "到手价"))
    wants_spec = any(k in question for k in ("规格", "等级", "重量", "都告诉", "都说一下", "有哪些"))
    order_intent = any(k in question for k in ("下单", "帮我订", "就要这个", "先来", "先拿", "来一箱", "来两箱", "怎么拍", "购买"))
    if order_intent and not (wants_price or wants_spec):
        return next_sales_step()
    sale_mode = detected_sale_mode()
    if (wants_price or wants_spec) and sale_mode == "整箱":
        rows = conn.execute(
            "SELECT product,price FROM price_rules WHERE sale_mode='整箱' AND status='启用' ORDER BY id"
        ).fetchall()
        selected_product = latest_choice(tuple(row["product"] for row in rows))
        selected = next((row for row in rows if row["product"] == selected_product), None)
        display_rows = [selected] if selected else rows
        price_text = "，".join(f"{row['product']} {yuan(row['price'])}元" for row in display_rows)
        answer = f"可以老板，{'这档' if selected else '整箱目前这几档'}：{price_text}。"
        region = latest_choice(("华南", "常规", "偏远"))
        freight_price = None
        if region:
            freight = conn.execute(
                "SELECT price FROM freight_rules WHERE region_group=? AND sale_mode='整箱' AND status='启用'",
                (region,),
            ).fetchone()
            if freight:
                freight_price = float(freight["price"])
                answer += f"您在{region}，整箱运费是{yuan(freight['price'])}元/箱。"
        else:
            answer += "您发我收货地区，我再把运费一起算上。"
        quantity_match = latest_regex(r"(\d+)\s*箱")
        if selected and quantity_match and freight_price is not None:
            quantity = int(quantity_match.group(1))
            goods_total = float(selected["price"]) * quantity
            total = goods_total + freight_price * quantity
            answer += f"按{quantity}箱算，商品{yuan(goods_total)}元，运费{yuan(freight_price * quantity)}元，合计{yuan(total)}元。"
        return answer, "整箱价格规则"

    if (wants_price or wants_spec) and sale_mode == "单粒":
        product = latest_choice(("猫山王", "黑刺"))
        if not product and latest_regex(r"(?:^|[^黑])刺"):
            product = "刺"
        grade_match = latest_regex(r"(?:\bAA\b|AA级|\bA\b|A级)", flags=re.IGNORECASE)
        grade = "AA" if grade_match and grade_match.group(0).upper().startswith("AA") else ("A" if grade_match else "")
        spec_match = latest_regex(r"\d+(?:\.\d+)?-\d+(?:\.\d+)?斤")
        if product and spec_match and (grade or product == "刺"):
            price_row = conn.execute(
                "SELECT price FROM price_rules WHERE sale_mode='单粒' AND product=? AND grade=? AND spec=? AND status='启用'",
                (product, grade, spec_match.group(0)),
            ).fetchone()
            if price_row:
                answer = f"可以老板，{product}{grade} {spec_match.group(0)}目前是{yuan(price_row['price'])}元/粒。"
                region = latest_choice(("华南", "常规", "偏远"))
                if region:
                    freight = conn.execute(
                        "SELECT price FROM freight_rules WHERE region_group=? AND sale_mode='单粒' AND status='启用'",
                        (region,),
                    ).fetchone()
                    if freight:
                        answer += f"{region}单粒运费是{yuan(freight['price'])}元/粒。"
                        quantity_match = latest_regex(r"(\d+)\s*粒")
                        if quantity_match:
                            quantity = int(quantity_match.group(1))
                            goods_total = float(price_row["price"]) * quantity
                            freight_total = float(freight["price"]) * quantity
                            answer += f"按{quantity}粒算，商品{yuan(goods_total)}元，运费{yuan(freight_total)}元，合计{yuan(goods_total + freight_total)}元。"
                else:
                    answer += "您把收货地区发我，我再把运费一起算上。"
                return answer, "单粒价格规则"
        return entry(
            "单粒价格承接-01",
            "单粒按品种、等级和重量段算，猫山王A/AA、黑刺A/AA以及6斤以上刺都有。表比较长，您先说想看哪一种，我把对应重量段和价格一次发清楚。",
        ), "单粒询价承接"

    if wants_price or wants_spec:
        return entry(
            "询价澄清-01",
            "可以老板，品种、规格和价格我都能发。整箱和单粒是两套表，您准备怎么拿货？我按对应的表给您发清楚。",
        ), "询价方式确认"

    if "运费" in question:
        return entry(
            "运费说明-01",
            "运费分华南、常规和偏远三档，整箱分别是60、120、180元，单粒分别是20、40、60元。您把收货地区发我，我帮您对一下档位。",
        ), "运费规则"
    if any(k in question for k in ("库存", "现货", "还有货", "有货吗", "能发多少")):
        return entry(
            "库存确认-01",
            "库存每天会随采购入库、订单占用和发货变化。我先按当前可发批次帮您核一下准确数量，确认后马上回复您。",
        ), "库存需人工核实"
    if any(k in question for k in ("少房", "少一房", "少两房", "虫果", "夹生", "生果", "核大", "售后", "退款", "赔偿")):
        return (
            "可以按已经确认的售后标准处理。请先把榴莲头和实际果况拍清楚，不强制要求完整开箱视频或快递面单；"
            "具体补偿金额和最终处理结果由售后人员结合订单与果况核实。",
            "必须人工确认",
        )
    if any(k in question for k in ("不需要", "别联系", "不要联系")):
        return "收到，抱歉打扰，我这边不再跟进，祝您生意兴隆。", "停止营销"
    if matches:
        safe = next((row for row in matches if row.get("risk_level") != "高"), matches[0])
        return safe["content"], safe.get("automation_level", "可自动建议")
    return entry(
        "未知信息承接-01",
        "这个我手上暂时没有准确信息，怕说错。您把具体想确认的内容发我，我帮您核实一下。",
    ), "知识不足，情境承接"


@app.get("/api/dashboard")
def dashboard():
    with db() as conn:
        lead_count = conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0]
        pending_tasks = conn.execute("SELECT COUNT(*) FROM friend_tasks WHERE status IN ('待执行','执行中')").fetchone()[0]
        accepted = conn.execute("SELECT COUNT(*) FROM friend_tasks WHERE status='已通过'").fetchone()[0]
        active_conversations = conn.execute("SELECT COUNT(*) FROM conversations").fetchone()[0]
        orders = conn.execute("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM orders").fetchone()
        low_stock = conn.execute(
            """SELECT COUNT(*) FROM products p
               LEFT JOIN product_inventory_settings s ON s.product_id=p.id
               WHERE p.status='在售' AND p.stock < COALESCE(s.reorder_point,10)"""
        ).fetchone()[0]
        pending_purchase = conn.execute(
            "SELECT COUNT(*) FROM purchase_orders WHERE status IN ('草稿','已审批','部分到货')"
        ).fetchone()[0]
        ready_to_ship = conn.execute(
            """SELECT COUNT(*) FROM orders o
               WHERE o.payment_status='已付款'
                 AND EXISTS (SELECT 1 FROM stock_reservations r WHERE r.order_id=o.id AND r.status='生效')"""
        ).fetchone()[0]
        shipment_exceptions = conn.execute(
            "SELECT COUNT(*) FROM fulfillment_shipments WHERE status='物流异常'"
        ).fetchone()[0]
        funnel = rows_to_dict(conn.execute("SELECT status name, COUNT(*) value FROM leads GROUP BY status ORDER BY value DESC").fetchall())
        recent = rows_to_dict(conn.execute("SELECT action,object_type,detail,created_at FROM audit_logs ORDER BY id DESC LIMIT 8").fetchall())
    return jsonify({
        "metrics": {
            "leads": lead_count,
            "pending_tasks": pending_tasks,
            "accepted": accepted,
            "conversations": active_conversations,
            "orders": orders[0],
            "revenue": orders[1],
            "low_stock": low_stock,
            "pending_purchase": pending_purchase,
            "ready_to_ship": ready_to_ship,
            "shipment_exceptions": shipment_exceptions,
        },
        "funnel": funnel,
        "recent": recent,
    })


@app.get("/api/knowledge/summary")
def knowledge_summary():
    with db() as conn:
        categories = rows_to_dict(conn.execute("SELECT category name,COUNT(*) count FROM knowledge_entries WHERE status='已发布' GROUP BY category ORDER BY count DESC").fetchall())
        return jsonify({
            "entries": conn.execute("SELECT COUNT(*) FROM knowledge_entries WHERE status='已发布'").fetchone()[0],
            "price_rules": conn.execute("SELECT COUNT(*) FROM price_rules WHERE status='启用'").fetchone()[0],
            "freight_rules": conn.execute("SELECT COUNT(*) FROM freight_rules WHERE status='启用'").fetchone()[0],
            "aftersales_rules": conn.execute("SELECT COUNT(*) FROM aftersales_rules WHERE status='启用'").fetchone()[0],
            "high_risk": conn.execute("SELECT COUNT(*) FROM knowledge_entries WHERE status='已发布' AND risk_level='高'").fetchone()[0],
            "categories": categories,
        })


@app.get("/api/knowledge/entries")
def list_knowledge_entries():
    category = request.args.get("category", "").strip()
    query = request.args.get("q", "").strip()
    with db() as conn:
        if query:
            data = search_knowledge_rows(conn, query, category, 200)
        else:
            sql = "SELECT * FROM knowledge_entries WHERE 1=1"
            params = []
            if category:
                sql += " AND category=?"
                params.append(category)
            sql += " ORDER BY category,title LIMIT 500"
            data = rows_to_dict(conn.execute(sql, params).fetchall())
    for row in data:
        if "structured_json" in row:
            try:
                row["structured"] = json.loads(row.pop("structured_json") or "{}")
            except json.JSONDecodeError:
                row["structured"] = {}
    return jsonify(data)


@app.post("/api/knowledge/entries")
def create_knowledge_entry():
    payload = request.get_json(force=True)
    content = (payload.get("content") or "").strip()
    title = (payload.get("title") or "").strip()
    if not title or not content:
        return jsonify({"error": "标题和内容不能为空"}), 400
    labels = auto_label_knowledge(content, (payload.get("category") or "").strip())
    with db() as conn:
        cur = conn.execute(
            """INSERT INTO knowledge_entries(category,title,content,structured_json,tags,risk_level,automation_level,status,source,version,created_at,updated_at)
               VALUES(?,?,?,?,?,?,?,?,?,1,?,?)""",
            (labels["category"], title, content, json.dumps(payload.get("structured") or {}, ensure_ascii=False),
             payload.get("tags") or labels["tags"], payload.get("risk_level") or labels["risk_level"],
             payload.get("automation_level") or labels["automation_level"], payload.get("status", "待审核"),
             payload.get("source", "控制台录入"), now(), now()),
        )
        audit(conn, "新增知识条目", "knowledge", cur.lastrowid, title)
    return jsonify({"id": cur.lastrowid, "labels": labels}), 201


@app.patch("/api/knowledge/entries/<int:entry_id>")
def update_knowledge_entry(entry_id: int):
    payload = request.get_json(force=True)
    allowed = {"category", "title", "content", "tags", "risk_level", "automation_level", "status", "source", "effective_from", "effective_to"}
    updates = {key: value for key, value in payload.items() if key in allowed}
    if not updates:
        return jsonify({"error": "没有可更新字段"}), 400
    updates["updated_at"] = now()
    with db() as conn:
        if not conn.execute("SELECT 1 FROM knowledge_entries WHERE id=?", (entry_id,)).fetchone():
            return jsonify({"error": "知识条目不存在"}), 404
        conn.execute(f"UPDATE knowledge_entries SET {','.join(f'{key}=?' for key in updates)} WHERE id=?", [*updates.values(), entry_id])
        audit(conn, "更新知识条目", "knowledge", entry_id, json.dumps(updates, ensure_ascii=False))
    return jsonify({"ok": True})


@app.get("/api/knowledge/rules")
def list_knowledge_rules():
    with db() as conn:
        return jsonify({
            "prices": rows_to_dict(conn.execute("SELECT * FROM price_rules WHERE status='启用' ORDER BY sale_mode,product,grade,spec").fetchall()),
            "freight": rows_to_dict(conn.execute("SELECT * FROM freight_rules WHERE status='启用' ORDER BY region_group,sale_mode").fetchall()),
            "aftersales": rows_to_dict(conn.execute("SELECT * FROM aftersales_rules WHERE status='启用' ORDER BY issue,id").fetchall()),
        })


@app.post("/api/knowledge/calculate")
def calculate_price_api():
    payload = request.get_json(force=True)
    try:
        with db() as conn:
            result = calculate_knowledge_price(conn, payload)
    except (ValueError, TypeError) as exc:
        return jsonify({"error": str(exc)}), 422
    return jsonify(result)


@app.post("/api/knowledge/answer")
def knowledge_answer_api():
    payload = request.get_json(force=True)
    question = (payload.get("question") or "").strip()
    if not question:
        return jsonify({"error": "请输入客户问题"}), 400
    with db() as conn:
        matches = search_knowledge_rows(conn, question, "", 12)
        context, requires_human, sources = build_answer_context(conn, question, matches)
        local_answer, decision_basis = local_knowledge_answer(conn, question, matches)
        requires_human = requires_human or decision_requires_human(decision_basis)
        setting = conn.execute("SELECT * FROM ai_settings WHERE id=1").fetchone()
    answer = local_answer
    mode = "本地规则"
    knowledge_hit = bool(matches or context)
    if not knowledge_hit:
        requires_human = requires_human or unknown_fact_requires_human(question)
    source_question = is_contact_source_question(question)
    sensitive_escalation = is_source_denial(question) or is_privacy_complaint(question) or is_stop_contact_request(question)
    automation_identity = is_automation_identity_question(question)
    protected_dialogue = source_question or sensitive_escalation or automation_identity
    if sensitive_escalation:
        requires_human = True
        knowledge_hit = True
    elif source_question:
        requires_human = False
        knowledge_hit = True
        sources = list(dict.fromkeys(["业务方确认-公开网站来源", *sources]))
    elif automation_identity:
        requires_human = False
        knowledge_hit = True
        sources = list(dict.fromkeys(["账号维护说明-01", *sources]))
    risk = knowledge_risk_level(matches, requires_human)
    if source_question and not sensitive_escalation:
        risk = "中"
    elif automation_identity:
        risk = "低"
    if payload.get("use_ai") is True and not protected_dialogue and setting and setting["status"] == "连接正常":
        api_key = reveal_api_key(setting["encrypted_api_key"])
        if api_key:
            system_prompt = (
                "你是榴莲B2B销售回复助手。知识库命中时，业务事实只能依据知识库回答。"
                "知识库未命中时，可以结合当前销售场景自然接话、回答低风险通用常识或追问澄清，但不得虚构商家、人员或商品事实。"
                "未命中时不得主动列举新的商品、品种、规格、服务或业务类型；只能使用客户已经提到的内容，或者提出不带预设选项的开放式问题。"
                "无论是否命中，都不得补充未提供的价格、库存、产地证明、时效或售后承诺。"
                "价格与运费信息不完整时必须追问；涉及退款、赔偿、食品质量争议时必须说明需要人工确认。"
                "回复要像真实销售，简短自然，一次只问一个主要问题，不使用虚假紧迫感。"
                "如果客户直接询问是否机器人、AI、真人或自动回复，必须如实说明账号由公司销售人员和服务系统共同维护，不得声称自己是真人或不是机器人。"
                f"\n当前问题知识库命中：{'是' if knowledge_hit else '否'}。\n知识库：\n" + (context or "无可靠知识")
            )
            try:
                response = deepseek_request(setting["base_url"], api_key, "/chat/completions", {
                    "model": setting["model"], "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": question}],
                    "temperature": 0.25, "max_tokens": 350, "stream": False,
                })
                answer_candidate = response["choices"][0]["message"]["content"].strip()
                if reply_contains_false_human_claim(answer_candidate):
                    raise RuntimeError("模型回复包含虚假真人声明")
                if not knowledge_hit and contextual_reply_has_unsupported_fact(answer_candidate):
                    raise RuntimeError("情境回复包含无知识支撑的业务事实")
                answer = answer_candidate
                mode = "DeepSeek+知识库" if knowledge_hit else "DeepSeek情境回复"
            except (RuntimeError, KeyError, IndexError) as exc:
                mode = "本地规则（AI连接失败）"
                decision_basis = f"{decision_basis}；AI连接失败：{str(exc)}"
    action = "暂停营销并转人工核查" if sensitive_escalation else ("人工确认后发送" if requires_human else "可自动建议")
    return jsonify({"answer": answer, "mode": mode, "risk": risk, "requires_human": requires_human,
                    "knowledge_hit": knowledge_hit, "decision_basis": decision_basis,
                    "action": action, "sources": sources, "matches": matches[:5]})


@app.get("/api/ai/settings")
def get_ai_settings():
    with db() as conn:
        row = conn.execute("SELECT * FROM ai_settings WHERE id=1").fetchone()
    data = dict(row)
    secret = reveal_api_key(data.pop("encrypted_api_key", ""))
    data["api_key_masked"] = f"{secret[:3]}****{secret[-4:]}" if len(secret) >= 8 else ("已配置" if secret else "")
    return jsonify(data)


@app.put("/api/ai/settings")
def save_ai_settings():
    payload = request.get_json(force=True)
    try:
        base_url = validate_deepseek_base_url(payload.get("base_url", "https://api.deepseek.com"))
        encrypted = protect_api_key((payload.get("api_key") or "").strip()) if payload.get("api_key") else None
    except (ValueError, RuntimeError) as exc:
        return jsonify({"error": str(exc)}), 400
    model = (payload.get("model") or "deepseek-v4-flash").strip()
    with db() as conn:
        if encrypted is None:
            conn.execute("UPDATE ai_settings SET base_url=?,model=?,updated_at=? WHERE id=1", (base_url, model, now()))
        else:
            conn.execute("UPDATE ai_settings SET base_url=?,model=?,encrypted_api_key=?,status='待测试',updated_at=? WHERE id=1", (base_url, model, encrypted, now()))
        audit(conn, "保存AI配置", "ai_settings", 1, f"DeepSeek / {model}，密钥已加密")
    return jsonify({"ok": True, "status": "待测试"})


@app.post("/api/ai/test")
def test_ai_settings():
    with db() as conn:
        setting = conn.execute("SELECT * FROM ai_settings WHERE id=1").fetchone()
        api_key = reveal_api_key(setting["encrypted_api_key"])
        if not api_key:
            return jsonify({"error": "请先保存API Key"}), 400
        try:
            response = deepseek_request(setting["base_url"], api_key, "/models")
            models = [item.get("id") for item in response.get("data", []) if item.get("id")]
            result = f"连接正常，可用模型{len(models)}个"
            conn.execute("UPDATE ai_settings SET status='连接正常',last_test_at=?,last_test_result=?,updated_at=? WHERE id=1", (now(), result, now()))
            audit(conn, "测试AI连接", "ai_settings", 1, result)
        except RuntimeError as exc:
            conn.execute("UPDATE ai_settings SET status='连接失败',last_test_at=?,last_test_result=?,updated_at=? WHERE id=1", (now(), str(exc), now()))
            return jsonify({"error": str(exc)}), 502
    return jsonify({"ok": True, "models": models, "result": result})


FIELD_ALIASES = {
    "门店名称": "store_name", "店名": "store_name", "名称": "store_name", "store_name": "store_name",
    "联系人": "contact_name", "联系人姓名": "contact_name", "contact_name": "contact_name",
    "手机号": "phone", "电话": "phone", "phone": "phone",
    "微信号": "wechat_id", "微信": "wechat_id", "wechat_id": "wechat_id",
    "地区": "region", "省市区": "region", "region": "region",
    "省份": "province", "province": "province",
    "城市": "city", "city": "city",
    "区/县": "district", "区县": "district", "省主区": "district", "district": "district",
    "门店类型": "store_type", "类型": "store_type", "转兰类型": "store_type", "store_type": "store_type",
    "来源": "source", "来源渠道": "source", "source": "source",
    "来源依据": "source_basis", "授权依据": "source_basis", "source_basis": "source_basis",
    "规模": "scale", "门店规模": "scale", "scale": "scale",
    "榴莲经营情况": "durian_status", "durian_status": "durian_status",
    "备注": "notes", "notes": "notes",
    "地址": "address", "address": "address",
    "经纬度": "coordinates", "坐标": "coordinates", "coordinates": "coordinates",
}

MAX_IMPORT_BYTES = 25 * 1024 * 1024
MAX_IMPORT_ROWS = 50000


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def canonical_field(header: Any) -> str:
    value = cell_text(header).replace("\ufeff", "").strip()
    return FIELD_ALIASES.get(value, FIELD_ALIASES.get(value.lower(), value))


def primary_phone(raw_phone: str) -> tuple[str, list[str]]:
    tokens = [item.strip() for item in re.split(r"[;；,，、|\s]+", raw_phone) if item.strip()]
    normalized = []
    for token in tokens:
        digits = re.sub(r"\D", "", token)
        if len(digits) == 13 and digits.startswith("86"):
            digits = digits[2:]
        if digits:
            normalized.append(digits)
    mobiles = [item for item in normalized if re.fullmatch(r"1[3-9]\d{9}", item)]
    return (mobiles[0] if mobiles else (normalized[0] if normalized else "")), normalized


def locate_header(rows: list[tuple[Any, ...]]) -> int:
    for index, values in enumerate(rows[:20]):
        fields = {canonical_field(value) for value in values if cell_text(value)}
        if "store_name" in fields and ({"phone", "wechat_id"} & fields):
            return index
    raise ValueError("没有找到包含门店名称和联系方式的表头")


def parse_uploaded_rows(raw: bytes, filename: str) -> tuple[list[tuple[str, int, dict[str, str]]], list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    parsed: list[tuple[str, int, dict[str, str]]] = []
    sheets: list[dict[str, Any]] = []
    if suffix == ".csv":
        text = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                pass
        if text is None:
            raise ValueError("CSV编码无法识别，请使用UTF-8或GBK")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;；")
        except csv.Error:
            dialect = csv.excel
        rows = list(csv.reader(io.StringIO(text), dialect=dialect))
        header_index = locate_header([tuple(row) for row in rows])
        headers = [canonical_field(value) for value in rows[header_index]]
        for row_no, values in enumerate(rows[header_index + 1:], start=header_index + 2):
            if not any(cell_text(value) for value in values):
                continue
            source_row = {headers[i]: cell_text(value) for i, value in enumerate(values) if i < len(headers) and headers[i]}
            parsed.append(("CSV", row_no, source_row))
            if len(parsed) > MAX_IMPORT_ROWS:
                raise ValueError(f"单次最多导入{MAX_IMPORT_ROWS}条数据")
        sheets.append({"name": "CSV", "rows": len(parsed), "headers": headers})
    elif suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
                if not rows or not any(any(cell_text(value) for value in row) for row in rows):
                    continue
                try:
                    header_index = locate_header(rows)
                except ValueError:
                    sheets.append({"name": worksheet.title, "rows": 0, "ignored": "未找到联系人表头"})
                    continue
                headers = [canonical_field(value) for value in rows[header_index]]
                sheet_count = 0
                for row_no, values in enumerate(rows[header_index + 1:], start=header_index + 2):
                    if not any(cell_text(value) for value in values):
                        continue
                    source_row = {headers[i]: cell_text(value) for i, value in enumerate(values) if i < len(headers) and headers[i]}
                    parsed.append((worksheet.title, row_no, source_row))
                    sheet_count += 1
                    if len(parsed) > MAX_IMPORT_ROWS:
                        raise ValueError(f"单次最多导入{MAX_IMPORT_ROWS}条数据")
                sheets.append({"name": worksheet.title, "rows": sheet_count, "headers": headers})
        finally:
            workbook.close()
    else:
        raise ValueError("仅支持.xlsx或.csv文件")
    if not parsed:
        raise ValueError("文件中没有可导入的数据行")
    return parsed, sheets


def is_wechat_searchable(wechat_id: str, phone: str) -> bool:
    return bool((wechat_id or "").strip() or re.fullmatch(r"1[3-9]\d{9}", (phone or "").strip()))


@app.post("/api/leads/import")
def import_leads():
    if "file" not in request.files:
        return jsonify({"error": "请选择Excel或CSV文件"}), 400
    upload = request.files["file"]
    filename = Path(upload.filename or "uploaded.csv").name
    raw = upload.read()
    if not raw:
        return jsonify({"error": "上传文件为空"}), 400
    if len(raw) > MAX_IMPORT_BYTES:
        return jsonify({"error": "文件超过25MB限制"}), 400
    try:
        source_rows, sheets = parse_uploaded_rows(raw, filename)
    except (ValueError, OSError, BadZipFile, InvalidFileException) as exc:
        return jsonify({"error": str(exc)}), 400
    imported = duplicate = invalid = rpa_ready = reference_only = unverified_source = 0
    errors = []
    with db() as conn:
        for sheet_name, line_no, source_row in source_rows:
            row = {canonical_field(key): cell_text(value) for key, value in source_row.items()}
            raw_phone = row.get("phone", "")
            phone, all_phones = primary_phone(raw_phone)
            row["phone"] = phone
            region_parts = [row.get("province", ""), row.get("city", ""), row.get("district", "")]
            row["region"] = row.get("region") or " ".join(dict.fromkeys(part for part in region_parts if part))
            provenance = f"{filename} / {sheet_name} / 第{line_no}行"
            row["source"] = row.get("source") or DEFAULT_CONTACT_SOURCE
            row["source_basis"] = row.get("source_basis") or DEFAULT_CONTACT_SOURCE_BASIS
            row["import_provenance"] = provenance
            note_parts = [row.get("notes", "")]
            if raw_phone and (len(all_phones) > 1 or raw_phone != phone):
                note_parts.append(f"原始电话：{raw_phone}")
            if row.get("address"):
                note_parts.append(f"地址：{row['address']}")
            if row.get("coordinates"):
                note_parts.append(f"经纬度：{row['coordinates']}")
            row["notes"] = "；".join(part for part in note_parts if part)
            if not row.get("store_name"):
                invalid += 1
                errors.append({"sheet": sheet_name, "line": line_no, "reason": "缺少门店名称"})
                continue
            if not row.get("phone") and not row.get("wechat_id"):
                invalid += 1
                errors.append({"sheet": sheet_name, "line": line_no, "reason": "没有可用电话号码或微信号"})
                continue
            try:
                conn.execute(
                    """INSERT INTO leads(store_name,contact_name,phone,wechat_id,region,store_type,source,source_basis,import_provenance,scale,durian_status,notes,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (row.get("store_name"), row.get("contact_name", ""), row.get("phone", ""), row.get("wechat_id", ""),
                     row.get("region", ""), row.get("store_type", ""), row.get("source", ""), row.get("source_basis", ""),
                     row.get("import_provenance", ""), row.get("scale", ""), row.get("durian_status", ""), row.get("notes", ""), now(), now()),
                )
                imported += 1
                if not contact_source_is_verified(row.get("source"), row.get("source_basis")):
                    unverified_source += 1
                if is_wechat_searchable(row.get("wechat_id", ""), row.get("phone", "")):
                    rpa_ready += 1
                else:
                    reference_only += 1
            except sqlite3.IntegrityError:
                duplicate += 1
        audit(conn, "导入线索", "lead", "batch", f"文件{filename}：成功{imported}条，来源待核验{unverified_source}条，重复{duplicate}条，无效{invalid}条")
    return jsonify({"filename": filename, "imported": imported, "duplicate": duplicate, "invalid": invalid,
                    "unverified_source": unverified_source, "rpa_ready": rpa_ready, "reference_only": reference_only,
                    "errors": errors[:50], "sheets": sheets})


@app.get("/api/leads")
def list_leads():
    status = request.args.get("status", "")
    keyword = request.args.get("q", "").strip()
    sql = "SELECT * FROM leads WHERE 1=1"
    params = []
    if status:
        sql += " AND status=?"
        params.append(status)
    if keyword:
        sql += " AND (store_name LIKE ? OR contact_name LIKE ? OR phone LIKE ? OR wechat_id LIKE ? OR region LIKE ?)"
        params.extend([f"%{keyword}%"] * 5)
    sql += " ORDER BY score DESC, id DESC LIMIT 500"
    with db() as conn:
        data = rows_to_dict(conn.execute(sql, params).fetchall())
    return jsonify(data)


@app.patch("/api/leads/<int:lead_id>")
def update_lead(lead_id: int):
    payload = request.get_json(force=True)
    allowed = {"status", "owner", "score", "tags", "notes", "next_follow_at", "stop_marketing"}
    updates = {k: v for k, v in payload.items() if k in allowed}
    if not updates:
        return jsonify({"error": "没有可更新字段"}), 400
    updates["updated_at"] = now()
    clauses = ",".join(f"{k}=?" for k in updates)
    with db() as conn:
        conn.execute(f"UPDATE leads SET {clauses} WHERE id=?", [*updates.values(), lead_id])
        audit(conn, "更新线索", "lead", lead_id, json.dumps(updates, ensure_ascii=False))
    return jsonify({"ok": True})


@app.get("/api/wechat/accounts")
def list_accounts():
    with db() as conn:
        reset_daily_quota(conn)
        return jsonify(rows_to_dict(conn.execute("SELECT * FROM wechat_accounts ORDER BY id").fetchall()))


@app.post("/api/wechat/accounts")
def create_account():
    p = request.get_json(force=True)
    with db() as conn:
        cur = conn.execute(
            "INSERT INTO wechat_accounts(nickname,wechat_no,channel_mode,approval_ref,daily_limit,quota_date,status,last_sync_at,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
            (p["nickname"], p["wechat_no"], "rpa", p.get("approval_ref", "腾讯RPA授权范围"), max(1, min(MAX_DAILY_FRIEND_ADDS, int(p.get("daily_limit", MAX_DAILY_FRIEND_ADDS)))), datetime.now().strftime("%Y-%m-%d"), "在线", now(), now()),
        )
        audit(conn, "新增微信账号", "wechat_account", cur.lastrowid, p["wechat_no"])
    return jsonify({"id": cur.lastrowid}), 201


@app.get("/api/friend-tasks")
def list_friend_tasks():
    with db() as conn:
        rows = conn.execute("""
            SELECT t.*, l.store_name,l.contact_name,l.wechat_id,l.phone,
                   COALESCE(NULLIF(t.region_snapshot,''),l.region) region,
                   l.source,l.source_basis,l.notes,
                   a.nickname account_name,a.wechat_no,
                   CASE
                     WHEN t.status IN ('已通过','已是好友') THEN '添加成功'
                     WHEN t.status IN ('未找到','执行失败') THEN '添加失败'
                     WHEN t.status='已发送' THEN '申请已发送'
                     WHEN t.status='需人工处理' THEN '暂停待处理'
                     ELSE t.status
                   END result_status
            FROM friend_tasks t JOIN leads l ON l.id=t.lead_id JOIN wechat_accounts a ON a.id=t.account_id
            ORDER BY t.id DESC LIMIT 500
        """).fetchall()
    return jsonify(rows_to_dict(rows))


@app.post("/api/friend-tasks")
def create_friend_tasks():
    p = request.get_json(force=True)
    lead_ids = [int(x) for x in p.get("lead_ids", [])]
    if not lead_ids:
        return jsonify({"error": "请选择线索"}), 400
    account_id = int(p["account_id"])
    greeting_template = p.get("greeting", "您好，我是榴莲产地供应链的销售，看到您在经营水果门店，想了解一下近期是否有榴莲采购计划？")
    remark_template = p.get("remark", "榴莲客户-{地区}-{店名}")
    created = skipped = unsearchable = unverified_source = 0
    with db() as conn:
        reset_daily_quota(conn)
        account = conn.execute("SELECT * FROM wechat_accounts WHERE id=?", (account_id,)).fetchone()
        if not account or account["status"] != "在线":
            return jsonify({"error": "微信账号不可用"}), 400
        remaining = max(0, min(MAX_DAILY_FRIEND_ADDS, account["daily_limit"]) - account["used_today"])
        for lead_id in lead_ids[:remaining]:
            lead = conn.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
            if not lead or lead["stop_marketing"]:
                skipped += 1
                continue
            if not contact_source_is_verified(lead["source"], lead["source_basis"]):
                unverified_source += 1
                continue
            if not is_wechat_searchable(lead["wechat_id"], lead["phone"]):
                unsearchable += 1
                continue
            exists = conn.execute("SELECT 1 FROM friend_tasks WHERE lead_id=? AND status IN ('待执行','执行中','已发送','已通过','已是好友')", (lead_id,)).fetchone()
            if exists:
                skipped += 1
                continue
            conn.execute(
                "INSERT INTO friend_tasks(lead_id,account_id,greeting,remark,region_snapshot,status,scheduled_at,created_at) VALUES(?,?,?,?,?,?,?,?)",
                (lead_id, account_id, render_task_text(greeting_template, lead, 50), render_task_text(remark_template, lead, 32), lead["region"] or "", "待执行", now(), now()),
            )
            created += 1
        audit(conn, "创建加好友任务", "friend_task", "batch", f"创建{created}条，来源待核验{unverified_source}条，跳过{skipped}条，无可搜索手机号{unsearchable}条")
    return jsonify({"created": created, "skipped": skipped, "unverified_source": unverified_source,
                    "unsearchable": unsearchable, "remaining_quota": remaining - created})


def execute_rpa_task(task_id: int) -> dict[str, Any]:
    with RPA_EXECUTION_LOCK:
        with db() as conn:
            reset_daily_quota(conn)
            task = conn.execute(
                """
                SELECT t.*,l.store_name,l.contact_name,l.wechat_id,l.phone,l.source,l.source_basis,
                       COALESCE(NULLIF(t.region_snapshot,''),l.region) region,
                       a.daily_limit,a.used_today,a.status account_status
                FROM friend_tasks t
                JOIN leads l ON l.id=t.lead_id
                JOIN wechat_accounts a ON a.id=t.account_id
                WHERE t.id=?
                """,
                (task_id,),
            ).fetchone()
            if not task:
                raise RPAError("任务不存在")
            if task["status"] != "待执行":
                raise RPAError(f"任务当前状态为“{task['status']}”，不能执行")
            if task["account_status"] != "在线":
                raise RPAError("执行账号不可用")
            if task["used_today"] >= min(MAX_DAILY_FRIEND_ADDS, task["daily_limit"]):
                raise RPAError("执行账号今日额度已用完")
            target = (task["wechat_id"] or task["phone"] or "").strip()
            if not target:
                raise RPAError("客户缺少微信号或手机号")
            if not contact_source_is_verified(task["source"], task["source_basis"]):
                raise RPAError("客户来源或授权依据尚未核验，禁止执行触达")
            task_data = dict(task)
            conn.execute("UPDATE friend_tasks SET status='执行中',result_note='正在调用微信RPA' WHERE id=?", (task_id,))
            audit(conn, "RPA开始执行", "friend_task", task_id, f"target={target}")

        try:
            result = WeixinDriver().submit_friend_request(
                target,
                task_data["greeting"],
                task_data.get("remark") or task_data["store_name"],
            )
        except Exception as exc:
            region_label = (task_data.get("region") or "地区未填写").strip()
            message = f"地区：{region_label}；添加失败：{str(exc)[:430]}"
            with db() as conn:
                conn.execute(
                    "UPDATE friend_tasks SET status='执行失败',executed_at=?,result_note=? WHERE id=?",
                    (now(), message, task_id),
                )
                audit(conn, "RPA执行失败", "friend_task", task_id, message)
            return {"task_id": task_id, "status": "执行失败", "result": "error", "message": message}

        with db() as conn:
            region_label = (task_data.get("region") or "地区未填写").strip()
            if result.status == SearchStatus.PENDING_VERIFICATION:
                status = "已发送"
                note = f"地区：{region_label}；RPA已提交申请，备注：{task_data.get('remark') or ''}，微信状态：等待验证"
                conn.execute("UPDATE wechat_accounts SET used_today=used_today+1,last_sync_at=? WHERE id=?", (now(), task_data["account_id"]))
                conn.execute("UPDATE leads SET status='已触达',last_contact_at=?,updated_at=? WHERE id=?", (now(), now(), task_data["lead_id"]))
            elif result.status == SearchStatus.ALREADY_FRIEND:
                status = "已是好友"
                note = f"地区：{region_label}；添加成功：RPA检测到已是好友，未重复发送"
                conn.execute("UPDATE leads SET status='已建立联系',updated_at=? WHERE id=?", (now(), task_data["lead_id"]))
            elif result.status == SearchStatus.NOT_FOUND:
                status = "未找到"
                note = f"地区：{region_label}；添加失败：RPA搜索无法找到该用户"
            elif result.status == SearchStatus.PAUSED:
                status = "需人工处理"
                note = f"地区：{region_label}；RPA已暂停：{result.message[:300]}"
            else:
                status = "需人工处理"
                note = f"地区：{region_label}；RPA返回未知结果，已停止后续操作"
            conn.execute(
                "UPDATE friend_tasks SET status=?,executed_at=?,result_note=? WHERE id=?",
                (status, now(), note, task_id),
            )
            audit(conn, "RPA结果回写", "friend_task", task_id, f"{status}：{note}")
        return {"task_id": task_id, "status": status, "result": result.status.value, "message": note}


@app.post("/api/friend-tasks/<int:task_id>/rpa-execute")
def execute_friend_task_by_rpa(task_id: int):
    try:
        result = execute_rpa_task(task_id)
    except RPAError as exc:
        return jsonify({"error": str(exc)}), 400
    return jsonify(result), (200 if result["status"] != "执行失败" else 400)


def rpa_job_snapshot(job_id: str) -> dict[str, Any]:
    with RPA_JOB_LOCK:
        job = RPA_JOBS.get(job_id)
        if not job:
            raise KeyError(job_id)
        snapshot = dict(job)
        snapshot["results"] = list(job["results"])
        return snapshot


def run_rpa_job(job_id: str, task_ids: list[int], interval_seconds: int):
    import pythoncom

    pythoncom.CoInitialize()
    try:
        with RPA_JOB_LOCK:
            RPA_JOBS[job_id]["status"] = "运行中"
            RPA_JOBS[job_id]["started_at"] = now()
            RPA_JOBS[job_id]["updated_at"] = now()
        for index, task_id in enumerate(task_ids):
            with RPA_JOB_LOCK:
                job = RPA_JOBS[job_id]
                if job["stop_requested"]:
                    job["status"] = "已停止"
                    job["updated_at"] = now()
                    break
                job["current_task_id"] = task_id
                job["updated_at"] = now()
            result = execute_rpa_task(task_id)
            with RPA_JOB_LOCK:
                job = RPA_JOBS[job_id]
                job["processed"] += 1
                job["results"].append(result)
                if result["status"] in ("已发送", "已是好友"):
                    job["success"] += 1
                else:
                    job["failed"] += 1
                job["updated_at"] = now()
                if result["status"] == "需人工处理":
                    job["status"] = "已暂停"
                    job["stop_reason"] = result["message"]
                    break
            if index < len(task_ids) - 1:
                time.sleep(interval_seconds)
        with RPA_JOB_LOCK:
            job = RPA_JOBS[job_id]
            if job["status"] == "运行中":
                job["status"] = "已完成"
            job["current_task_id"] = None
            job["finished_at"] = now()
            job["updated_at"] = now()
    except Exception as exc:
        with RPA_JOB_LOCK:
            job = RPA_JOBS[job_id]
            job["status"] = "执行失败"
            job["stop_reason"] = str(exc)[:500]
            job["current_task_id"] = None
            job["finished_at"] = now()
            job["updated_at"] = now()
    finally:
        pythoncom.CoUninitialize()


@app.post("/api/rpa/jobs")
def create_rpa_job():
    global RPA_LATEST_JOB_ID
    payload = request.get_json(force=True)
    account_id = int(payload["account_id"])
    interval_seconds = max(5, min(60, int(payload.get("interval_seconds", 8))))
    with RPA_JOB_LOCK:
        active = next((job for job in RPA_JOBS.values() if job["status"] in ("等待启动", "运行中")), None)
        if active:
            return jsonify({"error": "已有RPA队列正在执行", "job_id": active["id"]}), 409
    with db() as conn:
        reset_daily_quota(conn)
        account = conn.execute("SELECT * FROM wechat_accounts WHERE id=?", (account_id,)).fetchone()
        if not account or account["status"] != "在线":
            return jsonify({"error": "执行账号不可用"}), 400
        remaining = max(0, min(MAX_DAILY_FRIEND_ADDS, account["daily_limit"]) - account["used_today"])
        rows = conn.execute(
            "SELECT id FROM friend_tasks WHERE account_id=? AND status='待执行' ORDER BY id LIMIT ?",
            (account_id, remaining),
        ).fetchall()
        task_ids = [row[0] for row in rows]
        if not task_ids:
            return jsonify({"error": "没有可执行的待处理任务"}), 400
        audit(conn, "创建RPA队列", "rpa_job", "pending", f"账号{account_id}，任务{len(task_ids)}条")
    job_id = uuid4().hex[:12]
    job = {
        "id": job_id,
        "status": "等待启动",
        "account_id": account_id,
        "total": len(task_ids),
        "processed": 0,
        "success": 0,
        "failed": 0,
        "current_task_id": None,
        "stop_requested": False,
        "stop_reason": "",
        "interval_seconds": interval_seconds,
        "results": [],
        "created_at": now(),
        "started_at": None,
        "finished_at": None,
        "updated_at": now(),
    }
    with RPA_JOB_LOCK:
        RPA_JOBS[job_id] = job
        RPA_LATEST_JOB_ID = job_id
    Thread(target=run_rpa_job, args=(job_id, task_ids, interval_seconds), daemon=True).start()
    return jsonify(rpa_job_snapshot(job_id)), 201


@app.get("/api/rpa/jobs/current")
def current_rpa_job():
    with RPA_JOB_LOCK:
        job_id = RPA_LATEST_JOB_ID
    if not job_id:
        return jsonify({"status": "未启动", "results": []})
    return jsonify(rpa_job_snapshot(job_id))


@app.post("/api/rpa/jobs/<job_id>/stop")
def stop_rpa_job(job_id: str):
    with RPA_JOB_LOCK:
        job = RPA_JOBS.get(job_id)
        if not job:
            return jsonify({"error": "RPA队列不存在"}), 404
        if job["status"] not in ("等待启动", "运行中"):
            return jsonify({"error": "队列已终止"}), 400
        job["stop_requested"] = True
        job["updated_at"] = now()
    return jsonify({"ok": True, "status": "正在停止"})


@app.post("/api/friend-tasks/<int:task_id>/confirm-sent")
def confirm_friend_task_sent(task_id: int):
    payload = request.get_json(silent=True) or {}
    if payload.get("confirmed") is not True:
        return jsonify({"error": "请确认已在微信中手动发送好友申请"}), 400
    with db() as conn:
        reset_daily_quota(conn)
        task = conn.execute("SELECT * FROM friend_tasks WHERE id=?", (task_id,)).fetchone()
        if not task or task["status"] != "待执行":
            return jsonify({"error": "任务状态不可确认"}), 400
        account = conn.execute("SELECT * FROM wechat_accounts WHERE id=?", (task["account_id"],)).fetchone()
        if account["used_today"] >= min(MAX_DAILY_FRIEND_ADDS, account["daily_limit"]):
            return jsonify({"error": "今日额度已用完"}), 400
        region_label = (task["region_snapshot"] or "地区未填写").strip()
        note = f"地区：{region_label}；操作员确认已在微信客户端手动发送"
        conn.execute("UPDATE friend_tasks SET status='已发送',executed_at=?,result_note=? WHERE id=?", (now(), note, task_id))
        conn.execute("UPDATE wechat_accounts SET used_today=used_today+1,last_sync_at=? WHERE id=?", (now(), task["account_id"]))
        conn.execute("UPDATE leads SET status='已触达',last_contact_at=?,updated_at=? WHERE id=?", (now(), now(), task["lead_id"]))
        audit(conn, "人工确认已发送", "friend_task", task_id, note)
    return jsonify({"ok": True, "status": "已发送"})


@app.post("/api/friend-tasks/<int:task_id>/skip")
def skip_friend_task(task_id: int):
    payload = request.get_json(silent=True) or {}
    reason = (payload.get("reason") or "操作员跳过").strip()[:200]
    with db() as conn:
        task = conn.execute("SELECT * FROM friend_tasks WHERE id=?", (task_id,)).fetchone()
        if not task or task["status"] != "待执行":
            return jsonify({"error": "任务状态不可跳过"}), 400
        conn.execute("UPDATE friend_tasks SET status='已跳过',executed_at=?,result_note=? WHERE id=?", (now(), reason, task_id))
        audit(conn, "跳过触达任务", "friend_task", task_id, reason)
    return jsonify({"ok": True, "status": "已跳过"})


@app.post("/api/friend-tasks/<int:task_id>/accept")
def accept_friend_task(task_id: int):
    with db() as conn:
        task = conn.execute("SELECT * FROM friend_tasks WHERE id=?", (task_id,)).fetchone()
        if not task or task["status"] != "已发送":
            return jsonify({"error": "任务状态不可更新"}), 400
        region_label = (task["region_snapshot"] or "地区未填写").strip()
        conn.execute(
            "UPDATE friend_tasks SET status='已通过',result_note=?,executed_at=COALESCE(executed_at,?) WHERE id=?",
            (f"地区：{region_label}；添加成功：客户已通过好友申请", now(), task_id),
        )
        conn.execute("UPDATE leads SET status='已建立联系',updated_at=? WHERE id=?", (now(), task["lead_id"]))
        conv = conn.execute("SELECT id FROM conversations WHERE lead_id=?", (task["lead_id"],)).fetchone()
        if conv:
            conv_id = conv[0]
        else:
            cur = conn.execute("INSERT INTO conversations(lead_id,updated_at) VALUES(?,?)", (task["lead_id"], now()))
            conv_id = cur.lastrowid
        conn.execute("INSERT INTO messages(conversation_id,sender,content,created_at) VALUES(?,?,?,?)", (conv_id, "sales", task["greeting"], now()))
        audit(conn, "好友通过", "friend_task", task_id, f"创建/更新会话{conv_id}")
    return jsonify({"ok": True, "conversation_id": conv_id})


@app.get("/api/conversations")
def list_conversations():
    with db() as conn:
        rows = conn.execute("""
            SELECT c.*,l.store_name,l.contact_name,l.region,l.status lead_status,
              (SELECT content FROM messages m WHERE m.conversation_id=c.id ORDER BY m.id DESC LIMIT 1) last_message
            FROM conversations c JOIN leads l ON l.id=c.lead_id ORDER BY c.updated_at DESC
        """).fetchall()
    return jsonify(rows_to_dict(rows))


@app.get("/api/conversations/<int:conversation_id>/messages")
def get_messages(conversation_id: int):
    with db() as conn:
        rows = conn.execute("SELECT * FROM messages WHERE conversation_id=? ORDER BY id", (conversation_id,)).fetchall()
        conn.execute("UPDATE conversations SET unread=0 WHERE id=?", (conversation_id,))
    return jsonify(rows_to_dict(rows))


@app.post("/api/conversations/<int:conversation_id>/messages")
def send_message(conversation_id: int):
    p = request.get_json(force=True)
    sender = p.get("sender", "sales")
    content = p.get("content", "").strip()
    if not content:
        return jsonify({"error": "消息不能为空"}), 400
    with db() as conn:
        cur = conn.execute("INSERT INTO messages(conversation_id,sender,content,created_at) VALUES(?,?,?,?)", (conversation_id, sender, content, now()))
        if is_stop_contact_request(content):
            intent = "停止联系"
        elif is_source_denial(content) or is_privacy_complaint(content):
            intent = "隐私与来源异议"
        elif is_contact_source_question(content):
            intent = "联系方式来源"
        elif any(k in content for k in ("价格", "多少钱", "报价")):
            intent = "询价"
        else:
            intent = "采购咨询"
        sentiment = "负面" if any(k in content for k in ("投诉", "太差", "退款", "赔偿")) or is_privacy_complaint(content) else "中性"
        takeover = 1 if sentiment == "负面" or is_source_denial(content) or is_stop_contact_request(content) or any(k in content for k in ("账期", "合同", "退款", "赔偿")) else 0
        conn.execute(
            "UPDATE conversations SET intent=?,sentiment=?,human_takeover=MAX(human_takeover,?),updated_at=? WHERE id=?",
            (intent, sentiment, takeover, now(), conversation_id),
        )
        if sender == "customer" and is_stop_contact_request(content):
            conn.execute(
                "UPDATE leads SET stop_marketing=1,status='停止联系',updated_at=? WHERE id=(SELECT lead_id FROM conversations WHERE id=?)",
                (now(), conversation_id),
            )
            conn.execute(
                "UPDATE chat_bindings SET auto_reply=0,last_error='客户要求停止联系',updated_at=? WHERE conversation_id=?",
                (now(), conversation_id),
            )
        audit(conn, "发送消息", "conversation", conversation_id, f"sender={sender}")
    return jsonify({"id": cur.lastrowid, "intent": intent, "human_takeover": bool(takeover)})


@app.post("/api/conversations/<int:conversation_id>/suggest")
def suggest_reply(conversation_id: int):
    with db() as conn:
        conv = conn.execute(
            """SELECT c.*,l.store_name,l.region,l.source,l.source_basis,l.import_provenance,
                      l.stop_marketing,l.id lead_id
               FROM conversations c JOIN leads l ON l.id=c.lead_id WHERE c.id=?""",
            (conversation_id,),
        ).fetchone()
        if not conv:
            return jsonify({"error": "会话不存在"}), 404
        history_rows = conn.execute(
            "SELECT sender,content FROM messages WHERE conversation_id=? ORDER BY id DESC LIMIT 12",
            (conversation_id,),
        ).fetchall()
        history_rows = list(reversed(history_rows))
        customer_messages = [row["content"] for row in history_rows if row["sender"] == "customer"]
        message = customer_messages[-1] if customer_messages else "您好"
        history = "\n".join(("客户" if row["sender"] == "customer" else "销售") + "：" + row["content"] for row in history_rows)
        retrieval_query = "\n".join(row["content"] for row in history_rows)
        matches = search_knowledge_rows(conn, retrieval_query, "", 12)
        context, _, history_sources = build_answer_context(conn, retrieval_query, matches)
        policy_rows = conn.execute(
            "SELECT title,content FROM knowledge_entries WHERE category='对话原则' AND status='已发布' ORDER BY id LIMIT 12"
        ).fetchall()
        if policy_rows:
            policy_context = "\n".join(f"[{row['title']}] {row['content']}" for row in policy_rows)
            context = policy_context + ("\n" + context if context else "")
            history_sources = [row["title"] for row in policy_rows] + history_sources
        current_matches = search_knowledge_rows(conn, message, "", 12)
        current_context, requires_human, current_sources = build_answer_context(conn, message, current_matches)
        knowledge_hit = bool(current_matches or current_context)
        if not knowledge_hit:
            requires_human = requires_human or unknown_fact_requires_human(message)
        sources = list(dict.fromkeys(current_sources + history_sources))
        source_question = is_contact_source_question(message)
        source_denied = is_source_denial(message)
        privacy_complaint = is_privacy_complaint(message)
        stop_request = is_stop_contact_request(message)
        automation_identity = is_automation_identity_question(message)
        protected_dialogue = source_question or source_denied or privacy_complaint or stop_request or automation_identity
        verified_source = contact_source_is_verified(conv["source"], conv["source_basis"])
        if source_question and not (source_denied or privacy_complaint or stop_request):
            requires_human = False
            knowledge_hit = True
            source_label = "客户来源记录" if verified_source else "业务方确认-公开网站来源"
            sources = list(dict.fromkeys([source_label, *sources]))
        elif protected_dialogue:
            requires_human = False if automation_identity and not (source_denied or privacy_complaint or stop_request) else True
            knowledge_hit = True
        suggestion, decision_basis = local_knowledge_answer(conn, message, current_matches, history, conv)
        requires_human = requires_human or decision_requires_human(decision_basis)
        risk = knowledge_risk_level(current_matches, requires_human)
        if source_question and not (source_denied or privacy_complaint or stop_request):
            risk = "中"
        elif protected_dialogue:
            risk = "低" if automation_identity and not (source_denied or privacy_complaint or stop_request) else "高"
        if not knowledge_hit and decision_basis == "知识不足，情境承接":
            suggestion, decision_basis = "好的老板，您想先了解品种、规格，还是价格？", "低风险情境回复"
        setting = conn.execute("SELECT * FROM ai_settings WHERE id=1").fetchone()
    mode = "本地知识库"
    if not protected_dialogue and setting and setting["status"] == "连接正常":
        api_key = reveal_api_key(setting["encrypted_api_key"])
        if api_key:
            prompt = (
                f"你正在给水果门店客户回复微信。客户门店：{conv['store_name']}，地区：{conv['region']}。"
                "结合知识库和最近对话生成1至2句自然回复，不要像客服机器人；一次只问一个主要问题。"
                "先观察客户最近三句话的长度、称呼和语气，再匹配相近的回复长度与口吻；话术示例只能作为表达素材，不能机械照抄。"
                "客户只回OK、好、嗯、可以或行，代表允许继续，必须根据最近对话推进一个尚未确认的字段，不能停住。"
                "如果客户当前问题没有命中知识库，允许结合最近对话和榴莲批发销售身份自然接话、回答低风险通用常识或追问澄清。"
                "但不得虚构公司、人员、商品、价格、库存、物流、证明或售后事实；不确定的业务事实要明确说需要确认。"
                "未命中时不得主动列举新的商品、品种、规格、服务或业务类型；只能承接客户已说内容，或者提出不带预设选项的开放式问题。"
                "先承接客户刚说的内容，不重复已经问过且客户已经回答的问题。成交前依次确认销售方式、品种、等级、重量规格、数量和地区档位。"
                "这里的销售方式只允许问整箱拿货还是单粒代发，不要问批发还是零售。"
                "只有知识库出现的价格与运费才能引用，不得编造库存、产地证明或售后承诺；信息不足就追问。"
                "不得使用保证新鲜、保证每颗品质一致、绝对没有问题等无法核验的承诺。"
                "退款赔偿只解释规则并说明需人工核实，但仍可继续普通沟通。\n"
                "如果客户直接询问是否机器人、AI、真人或自动回复，必须如实说明账号由公司销售人员和服务系统共同维护，不得声称自己是真人或不是机器人。\n"
                f"客户当前问题知识库命中：{'是' if knowledge_hit else '否'}。\n"
                "最近对话：\n" + history + "\n知识库：\n" + (context or "无可靠知识")
            )
            try:
                response = deepseek_request(setting["base_url"], api_key, "/chat/completions", {
                    "model": setting["model"], "messages": [{"role": "system", "content": prompt}, {"role": "user", "content": message}],
                    "temperature": 0.25, "max_tokens": 220, "stream": False,
                })
                choice = response["choices"][0]
                suggestion_candidate = choice["message"]["content"].strip()
                if reply_contains_false_human_claim(suggestion_candidate):
                    raise RuntimeError("模型回复包含虚假真人声明")
                if choice.get("finish_reason") == "length" or not suggestion_candidate:
                    raise RuntimeError("模型回复被截断或为空")
                if suggestion_candidate[-1] not in "。！？!?）)🤝😊":
                    raise RuntimeError("模型回复不是完整句子")
                if not knowledge_hit and contextual_reply_has_unsupported_fact(suggestion_candidate):
                    raise RuntimeError("情境回复包含无知识支撑的业务事实")
                suggestion = suggestion_candidate
                mode = "DeepSeek+知识库" if knowledge_hit else "DeepSeek情境回复"
            except (RuntimeError, KeyError, IndexError):
                mode = "本地知识库（AI不可用）"
    action = "暂停营销并转人工核查" if requires_human and protected_dialogue else ("可自动建议" if not requires_human else "人工确认后发送")
    return jsonify({"suggestion": suggestion, "risk": risk, "basis": mode, "sources": sources,
                    "decision_basis": decision_basis, "action": action,
                    "requires_human": requires_human, "knowledge_hit": knowledge_hit})


@app.get("/api/products")
def list_products():
    with db() as conn:
        return jsonify(rows_to_dict(conn.execute("SELECT * FROM products ORDER BY id").fetchall()))


@app.post("/api/quotes")
def create_quote():
    p = request.get_json(force=True)
    with db() as conn:
        product = conn.execute("SELECT * FROM products WHERE id=?", (int(p["product_id"]),)).fetchone()
        quantity = int(p["quantity"])
        unit_price = float(p.get("unit_price", product["price"]))
        freight = float(p.get("freight", 0))
        total = round(quantity * unit_price + freight, 2)
        valid_until = (datetime.now() + timedelta(days=int(p.get("valid_days", 3)))).strftime("%Y-%m-%d")
        cur = conn.execute("INSERT INTO quotes(lead_id,product_id,quantity,unit_price,freight,total,valid_until,status,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                           (int(p["lead_id"]), product["id"], quantity, unit_price, freight, total, valid_until, "待确认", now()))
        conn.execute("UPDATE leads SET status='报价中',updated_at=? WHERE id=?", (now(), int(p["lead_id"])))
        audit(conn, "创建报价", "quote", cur.lastrowid, f"{product['name']} x {quantity} = {total}")
    return jsonify({"id": cur.lastrowid, "total": total, "valid_until": valid_until}), 201


@app.get("/api/quotes")
def list_quotes():
    with db() as conn:
        rows = conn.execute("SELECT q.*,l.store_name,p.name product_name,p.unit FROM quotes q JOIN leads l ON l.id=q.lead_id JOIN products p ON p.id=q.product_id ORDER BY q.id DESC").fetchall()
    return jsonify(rows_to_dict(rows))


@app.post("/api/orders")
def create_order():
    p = request.get_json(force=True)
    with db() as conn:
        quote = conn.execute("SELECT * FROM quotes WHERE id=?", (int(p["quote_id"]),)).fetchone()
        if not quote:
            return jsonify({"error": "报价不存在"}), 404
        order_no = "DR" + datetime.now().strftime("%Y%m%d%H%M%S") + f"{quote['id']:03d}"
        warehouse = conn.execute("SELECT id FROM warehouses WHERE is_default=1 AND status='启用' ORDER BY id LIMIT 1").fetchone()
        if not warehouse:
            return jsonify({"error": "没有启用的默认仓库"}), 409
        cur = conn.execute(
            """
            INSERT INTO orders(order_no,lead_id,quote_id,amount,payment_status,status,receiver,phone,address,
              warehouse_id,inventory_status,fulfillment_status,sales_channel,created_at,updated_at)
            VALUES(?,?,?,?,?,'待占库',?,?,?,?, '未占库','未发货','微信',?,?)
            """,
            (order_no, quote["lead_id"], quote["id"], quote["total"], p.get("payment_status", "待付款"),
             p.get("receiver", ""), p.get("phone", ""), p.get("address", ""), warehouse["id"], now(), now()),
        )
        ensure_order_item_from_quote(conn, int(cur.lastrowid), quote, now())
        conn.execute("UPDATE quotes SET status='已转订单' WHERE id=?", (quote["id"],))
        conn.execute("UPDATE leads SET status='履约中',updated_at=? WHERE id=?", (now(), quote["lead_id"]))
        audit(conn, "创建订单", "order", cur.lastrowid, order_no)
    return jsonify({"id": cur.lastrowid, "order_no": order_no}), 201


@app.get("/api/orders")
def list_orders():
    with db() as conn:
        rows = conn.execute("""
          SELECT o.*,l.store_name,w.name warehouse_name,
            (SELECT GROUP_CONCAT(p.name||' ×'||i.quantity,'；') FROM order_items i JOIN products p ON p.id=i.product_id WHERE i.order_id=o.id) item_summary,
            COALESCE((SELECT SUM(i.quantity) FROM order_items i WHERE i.order_id=o.id),0) ordered_quantity,
            COALESCE((SELECT SUM(i.fulfilled_qty) FROM order_items i WHERE i.order_id=o.id),0) fulfilled_quantity,
            COALESCE((SELECT SUM(r.quantity) FROM stock_reservations r WHERE r.order_id=o.id AND r.status='生效'),0) reserved_quantity,
            (SELECT s.carrier FROM fulfillment_shipments s WHERE s.order_id=o.id ORDER BY s.id DESC LIMIT 1) carrier,
            (SELECT s.tracking_no FROM fulfillment_shipments s WHERE s.order_id=o.id ORDER BY s.id DESC LIMIT 1) tracking_no,
            (SELECT s.status FROM fulfillment_shipments s WHERE s.order_id=o.id ORDER BY s.id DESC LIMIT 1) shipment_status,
            (SELECT COUNT(*) FROM fulfillment_shipments s WHERE s.order_id=o.id) shipment_count
          FROM orders o JOIN leads l ON l.id=o.lead_id LEFT JOIN warehouses w ON w.id=o.warehouse_id
          ORDER BY o.id DESC
        """).fetchall()
    return jsonify(rows_to_dict(rows))


@app.patch("/api/orders/<int:order_id>")
def update_order(order_id: int):
    p = request.get_json(force=True)
    allowed = {"payment_status", "receiver", "phone", "address"}
    updates = {k: v for k, v in p.items() if k in allowed}
    if "payment_status" in updates and updates["payment_status"] not in {"待付款", "已付款"}:
        return jsonify({"error": "支付状态无效"}), 400
    updates["updated_at"] = now()
    with db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return jsonify({"error": "订单不存在"}), 404
        if updates.get("payment_status") == "已付款":
            updates["status"] = "待发货" if order["inventory_status"] == "已占库" else "待占库"
        conn.execute(f"UPDATE orders SET {','.join(f'{k}=?' for k in updates)} WHERE id=?", [*updates.values(), order_id])
        audit(conn, "更新订单", "order", order_id, json.dumps(updates, ensure_ascii=False))
    return jsonify({"ok": True})


@app.post("/api/shipments")
def create_shipment():
    p = request.get_json(force=True)
    with db() as conn:
        conn.execute("BEGIN IMMEDIATE")
        try:
            result = create_fulfillment_shipment(conn, p, now(), audit)
        except InventoryError as exc:
            return jsonify({"error": str(exc)}), exc.status_code
        except sqlite3.IntegrityError:
            return jsonify({"error": "运单号已经存在，禁止重复发货"}), 409
    return jsonify(result), 201


@app.get("/api/audit-logs")
def list_audit_logs():
    with db() as conn:
        return jsonify(rows_to_dict(conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 200").fetchall()))


def build_operations_suggestion(conversation_id: int) -> dict[str, Any] | None:
    """Reuse the same guarded knowledge/DeepSeek path for worker-generated replies."""
    response = suggest_reply(conversation_id)
    if isinstance(response, tuple):
        return None
    return response.get_json()


register_inventory_routes(app, db, now, audit)
register_operations_routes(app, db, now, audit, build_operations_suggestion)


if __name__ == "__main__":
    init_db()
    app.run(host="127.0.0.1", port=8015, debug=False)
